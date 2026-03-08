"""
Secure Dealer Vehicle & Key Tracker  ·  Ultimate Edition
Flask Backend — Airtable · OpenAI · NHTSA · All Premium Features
"""

import os, re, json, csv, io, uuid, secrets
import requests
try:
    import qrcode
    from qrcode.image.pure import PyPNGImage as _PyPNG
    HAS_QRCODE = True
except ImportError:
    HAS_QRCODE = False
from datetime import datetime, timezone
from functools import wraps
from collections import Counter, defaultdict
from werkzeug.security import generate_password_hash, check_password_hash
from flask import (
    Flask, render_template, request,
    session, redirect, url_for, jsonify, abort,
    make_response
)
from openai import OpenAI

app = Flask(__name__)
_secret = os.environ.get("FLASK_SECRET_KEY")
if not _secret:
    raise RuntimeError("FLASK_SECRET_KEY environment variable must be set")
app.secret_key = _secret

# ── Config ─────────────────────────────────────────────────────────────────────
AIRTABLE_TOKEN   = os.environ.get("AIRTABLE_TOKEN", "")
AIRTABLE_BASE_ID = os.environ.get("AIRTABLE_BASE_ID", "")
AIRTABLE_TABLE   = os.environ.get("AIRTABLE_TABLE", "Vehicles")
OPENAI_API_KEY   = os.environ.get("OPENAI_API_KEY", "")
ACCESS_PASSWORD  = os.environ.get("ACCESS_PASSWORD", "dealer2024")
ADMIN_PASSWORD   = os.environ.get("ADMIN_PASSWORD",  "admin2024")
URL_SECRET_TOKEN = os.environ.get("URL_SECRET_TOKEN", "")
DEALERSHIP_NAME  = os.environ.get("DEALERSHIP_NAME", "Premier Auto Group")
KEY_OVERDUE_HOURS = int(os.environ.get("KEY_OVERDUE_HOURS", "1"))
MONTHLY_TARGET    = int(os.environ.get("MONTHLY_TARGET", "15"))
# In-memory overrides (survive until server restart, settable from UI)
_settings = {
    "key_overdue_hours": KEY_OVERDUE_HOURS,  # default 1h
    "monthly_target":    MONTHLY_TARGET,
    "dealership_name":   DEALERSHIP_NAME,
}
TWILIO_SID       = os.environ.get("TWILIO_SID", "")
TWILIO_TOKEN_ENV = os.environ.get("TWILIO_AUTH_TOKEN", "")
TWILIO_FROM      = os.environ.get("TWILIO_FROM", "")
MANAGER_PHONE    = os.environ.get("MANAGER_PHONE", "")
DEALERSHIP_PHONE = os.environ.get("DEALERSHIP_PHONE", MANAGER_PHONE)  # public-facing phone
DEALERSHIP_ADDRESS = os.environ.get("DEALERSHIP_ADDRESS", "")          # e.g. 149 Terminus St Liverpool NSW

VEHICLE_STATUSES = ["Available","Sold","In Service","Demo","Hold","Pending","Archived"]
VEHICLE_COLOURS = [
    "White","Pearl White","White Metallic",
    "Silver","Silver Metallic",
    "Grey","Grey Metallic","Gunmetal",
    "Black","Black Metallic",
    "Blue","Blue Metallic","Navy","Sky Blue",
    "Red","Red Metallic","Burgundy",
    "Orange","Yellow","Green","Green Metallic",
    "Brown","Brown Metallic","Beige","Champagne",
    "Gold","Bronze","Titanium","Purple",
    "Two-Tone","Other",
]

def clean_status(raw):
    """Strip ALL surrounding quotes/whitespace from a status value."""
    if not raw: return ""
    s = str(raw).strip()
    s = s.strip('"').strip("'").strip()
    s = s.strip('"').strip("'").strip()
    return s

NHTSA_VIN_URL    = "https://vpic.nhtsa.dot.gov/api/vehicles/decodevinvalues/{vin}?format=json"
NHTSA_RECALL_URL = "https://api.nhtsa.gov/recalls/recallsByVehicle?make={make}&model={model}&modelYear={year}"

openai_client = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

# ── In-memory stores ───────────────────────────────────────────────────────────
_activity_log  = []
_audit_trail   = []
_enquiries     = []
_test_drives   = []
_service_log   = []
_key_checkouts = {}
_pipeline      = []   # deal pipeline cards
_trade_ins     = []   # trade-in vehicles       # tag_id -> {staff, time, vehicle}
_price_history = defaultdict(list)  # internal_id -> [{price, date, user}]
_staff_accounts = {
    "staff":   {"name": "Staff",  "role": "staff",  "password": generate_password_hash(ACCESS_PASSWORD), "phone": "", "email": ""},
    "admin":   {"name": "Admin",  "role": "admin",  "password": generate_password_hash(ADMIN_PASSWORD),  "phone": "", "email": ""},
    # Add more accounts via Admin → Staff. Each staff member has their own phone number.
    # When a staff member shares a link (/share/<id>?ref=username) the customer sees THEIR number.
    # Admin can edit any staff phone at any time — if someone loses their phone, just update the number.
}

# ── Logging ────────────────────────────────────────────────────────────────────

def log_activity(action, internal_id="", detail="", user="staff"):
    _activity_log.insert(0, {
        "id": str(uuid.uuid4())[:8],
        "timestamp": datetime.now(timezone.utc).strftime("%b %d %H:%M"),
        "timestamp_iso": datetime.now(timezone.utc).isoformat(),
        "action": action, "internal_id": internal_id,
        "detail": detail, "user": user,
    })
    if len(_activity_log) > 500:
        _activity_log.pop()

def log_audit(action, internal_id, field, old_val, new_val, user):
    _audit_trail.insert(0, {
        "timestamp": datetime.now(timezone.utc).strftime("%b %d %H:%M UTC"),
        "timestamp_iso": datetime.now(timezone.utc).isoformat(),
        "action": action, "internal_id": internal_id,
        "field": field, "old": str(old_val or ""), "new": str(new_val or ""), "user": user,
    })
    if len(_audit_trail) > 1000:
        _audit_trail.pop()

# ── Airtable ───────────────────────────────────────────────────────────────────

def at_headers():
    return {"Authorization": f"Bearer {AIRTABLE_TOKEN}", "Content-Type": "application/json"}

def at_url():
    return f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{AIRTABLE_TABLE}"

def get_vehicle(internal_id):
    url = f"{at_url()}?filterByFormula={{Internal_ID}}='{internal_id}'&maxRecords=1"
    r = requests.get(url, headers=at_headers(), timeout=10)
    if r.status_code != 200: return None
    recs = r.json().get("records", [])
    if not recs: return None
    fields = recs[0].get("fields", {})
    if fields.get("Status"):
        fields["Status"] = clean_status(fields["Status"])
    return fields

def get_record_id(internal_id):
    url = f"{at_url()}?filterByFormula={{Internal_ID}}='{internal_id}'&maxRecords=1"
    r = requests.get(url, headers=at_headers(), timeout=10)
    if r.status_code != 200: return None
    recs = r.json().get("records", [])
    return recs[0]["id"] if recs else None

def get_all_vehicles(include_archived=False):
    """Paginated fetch — no 200-record cap."""
    filt = "" if include_archived else "?filterByFormula=NOT({Status}='Archived')"
    sep  = "&" if filt else "?"
    base = f"{at_url()}{filt}{sep}sort[0][field]=Internal_ID&sort[0][direction]=asc&pageSize=100"
    out, offset = [], None
    while True:
        url = base + (f"&offset={offset}" if offset else "")
        try:
            r = requests.get(url, headers=at_headers(), timeout=15)
            if r.status_code != 200:
                break
            data = r.json()
            for rec in data.get("records", []):
                flds = rec.get("fields", {})
                flds["_record_id"] = rec["id"]
                if flds.get("Status"):
                    flds["Status"] = clean_status(flds["Status"])
                if flds.get("Date_Added"):
                    try:
                        added = datetime.fromisoformat(flds["Date_Added"].replace("Z",""))
                        if added.tzinfo is None:
                            added = added.replace(tzinfo=timezone.utc)
                        flds["_days_on_lot"] = (datetime.now(timezone.utc) - added).days
                        flds["_dot_flag"] = dot_flag(flds["_days_on_lot"])
                    except Exception:
                        flds["_days_on_lot"] = None; flds["_dot_flag"] = ""
                else:
                    flds["_days_on_lot"] = None; flds["_dot_flag"] = ""
                flds["_profit"] = calc_profit(flds)
                out.append(flds)
            offset = data.get("offset")
            if not offset:
                break
        except Exception:
            break
    return out

def patch_vehicle(internal_id, fields, user="system", audit=True):
    rid = get_record_id(internal_id)
    if not rid: return False, "Not found"
    if audit and user not in ("system",):
        cur = get_vehicle(internal_id) or {}
        for k, v in fields.items():
            old = cur.get(k, "")
            if str(old) != str(v or ""):
                log_audit("field_change", internal_id, k, old or "—", v or "—", user)
    select_fields = {"Status"}
    safe_fields = {
        k: (clean_status(v) if k in select_fields and v else v)
        for k, v in fields.items()
    }
    url = f"{at_url()}/{rid}"
    r = requests.patch(url, headers=at_headers(), json={"fields": safe_fields}, timeout=10)
    if r.status_code == 200:
        return True, "OK"
    # If Airtable rejects due to unknown field, strip offending fields and retry
    err_text = r.text
    if r.status_code == 422 and "UNKNOWN_FIELD_NAME" in err_text:
        try:
            bad_field = re.search(r'"Unknown field name: \"([^\"]+)\"', err_text)
            if not bad_field:
                bad_field = re.search(r'Unknown field name: "([^"]+)"', err_text)
            if bad_field:
                bad = bad_field.group(1)
                safe_fields.pop(bad, None)
                r2 = requests.patch(url, headers=at_headers(), json={"fields": safe_fields}, timeout=10)
                if r2.status_code == 200:
                    return True, f"OK (skipped unknown field: {bad})"
                return False, r2.text[:300]
        except Exception:
            pass
    return False, err_text[:300]

def create_vehicle_record(fields):
    if get_record_id(fields.get("Internal_ID","")):
        return False, f"ID {fields['Internal_ID']} already exists"
    if not fields.get("Date_Added"):
        fields["Date_Added"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    r = requests.post(at_url(), headers=at_headers(), json={"fields": fields}, timeout=10)
    for _attempt in range(10):
        if r.status_code in (200, 201):
            return True, "Created"
        if r.status_code != 422:
            break
        err = r.text
        bad = None
        try:
            if "UNKNOWN_FIELD_NAME" in err:
                m = re.search(r'Unknown field name: "([^"]+)"', err) or re.search(r'Unknown field name: \\"([^\\"]+)\\"', err)
                if m: bad = m.group(1); fields.pop(bad, None)
            elif "INVALID_VALUE_FOR_COLUMN" in err:
                m = re.search(r'Field \\"([^\\"]+)\\" cannot accept', err) or re.search(r'Cannot parse value [^f]+for field (\w+)', err)
                if m:
                    bad = m.group(1)
                    conv = _to_num(str(fields.get(bad,"")))
                    fields[bad] = conv if conv is not None else fields.pop(bad, None)
        except Exception:
            break
        if not bad: break
        r = requests.post(at_url(), headers=at_headers(), json={"fields": fields}, timeout=10)
    return False, r.text[:200]
def delete_vehicle_record(internal_id):
    rid = get_record_id(internal_id)
    if not rid: return False, "Not found"
    r = requests.delete(f"{at_url()}/{rid}", headers=at_headers(), timeout=10)
    return (True, "Deleted") if r.status_code == 200 else (False, r.text[:200])

def free_tags(all_v):
    return [{"tag": v["Key_Tag_ID"], "was": v.get("Internal_ID","?")}
            for v in all_v if v.get("Status") in ("Sold","Archived") and v.get("Key_Tag_ID")]

def next_id(all_v, prefix="CAR"):
    pat = re.compile(rf"^{re.escape(prefix)}-(\d+)$", re.IGNORECASE)
    nums = [int(m.group(1)) for v in all_v if (m := pat.match(v.get("Internal_ID","")))]
    return f"{prefix}-{(max(nums)+1 if nums else 1):02d}"

def _to_num(val, as_int=True):
    """Convert any human-entered number to int/float for Airtable Number fields.
    Handles: "$25,000", "50,000 km", "50k", "12423", "25000.00" etc.
    """
    if val is None: return None
    s = str(val).strip().lower()
    if not s: return None
    # Handle "k" suffix: "50k" -> 50000
    multiplier = 1
    if s.endswith("k"):
        multiplier = 1000
        s = s[:-1]
    # Strip everything except digits and decimal point
    s = re.sub(r"[^\d.]", "", s)
    if not s: return None
    try:
        f = float(s) * multiplier
        return int(f) if as_int else f
    except ValueError:
        return None

def form_to_fields(form, include_id=True):
    out = {}
    if include_id:
        out["Internal_ID"] = form.get("internal_id","").strip().upper()
    out.update({
        "Make":          form.get("make","").strip(),
        "Model":         form.get("model","").strip(),
        "Year":          int(form.get("year",0) or 0) or None,
        "Color":         form.get("color","").strip(),
        "VIN":           form.get("vin","").strip().upper(),
        "Mileage":       _to_num(form.get("mileage","")) or form.get("mileage","").strip(),
        "Price":         _to_num(form.get("price","")),
        "Trim":          form.get("trim","").strip(),
        "Engine":        form.get("engine","").strip(),
        "Status":        clean_status(form.get("status","Available")) or "Available",
        "Key_Tag_ID":    form.get("key_tag_id","").strip(),
        "Deep_Link_URL": form.get("deep_link_url","").strip(),
        "Notes":         form.get("notes","").strip(),
        "Purchase_Price":_to_num(form.get("purchase_price","")),
        "Recon_Cost":    _to_num(form.get("recon_cost","")),
        "Transmission":  form.get("transmission","").strip(),
        "Body":          form.get("body","").strip(),
        "Drive":         form.get("drive","").strip(),
        "Fuel":          form.get("fuel","").strip(),
    })
    return {k: v for k, v in out.items() if v not in (None, "", 0)}

# ── Helpers ────────────────────────────────────────────────────────────────────

def dot_flag(days):
    if days is None: return ""
    if days >= 90: return "dot-red"
    if days >= 60: return "dot-amber"
    if days >= 30: return "dot-yellow"
    return "dot-green"

def overdue_keys():
    """Return list of key checkouts that have been out longer than threshold."""
    threshold_hours = _settings.get("key_overdue_hours", 2)
    now = datetime.now(timezone.utc)
    overdue = []
    for tag, co in _key_checkouts.items():
        try:
            checked_out = datetime.fromisoformat(co["timestamp"])
            if checked_out.tzinfo is None:
                checked_out = checked_out.replace(tzinfo=timezone.utc)
            hours_out = (now - checked_out).total_seconds() / 3600
            if hours_out >= threshold_hours:
                overdue.append({**co, "hours_out": round(hours_out, 1)})
        except Exception:
            pass
    return sorted(overdue, key=lambda x: x.get("hours_out", 0), reverse=True)

def sold_this_month(vehicles):
    """Count vehicles sold in the current calendar month."""
    now = datetime.now()
    count = 0
    for v in vehicles:
        if v.get("Status") != "Sold": continue
        sd = v.get("Sold_Date","")
        if sd:
            try:
                d = datetime.fromisoformat(sd[:10])
                if d.year == now.year and d.month == now.month:
                    count += 1
            except Exception:
                pass
    return count

def due_followups():
    """Return enquiries with a follow_up_date of today or earlier that are still pending."""
    today = datetime.now().strftime("%Y-%m-%d")
    return [e for e in _enquiries
            if e.get("follow_up_date") and e["follow_up_date"] <= today
            and e.get("follow_up_status") == "pending"]


def health_score(vehicle):
    """Return 0-100 health score + breakdown."""
    score = 100
    issues = []
    bonuses = []
    days = vehicle.get("_days_on_lot")
    price_raw = re.sub(r"[^\d.]", "", str(vehicle.get("Price","") or ""))
    purchase_raw = re.sub(r"[^\d.]", "", str(vehicle.get("Purchase_Price","") or ""))
    mileage_raw = re.sub(r"[^\d]", "", str(vehicle.get("Mileage","") or ""))
    try: price = float(price_raw)
    except: price = 0
    try: purchase = float(purchase_raw)
    except: purchase = 0
    try: mileage = int(mileage_raw)
    except: mileage = 0
    if days is not None:
        if days > 90:   score -= 25; issues.append("On lot " + str(days) + " days — very slow mover")
        elif days > 60: score -= 15; issues.append("On lot " + str(days) + " days — needs attention")
        elif days > 30: score -= 5;  issues.append("On lot " + str(days) + " days")
        else:           bonuses.append("Fresh stock (" + str(days) + " days)")
    if mileage > 200000: score -= 15; issues.append("Very high mileage (" + str(mileage) + "km)")
    elif mileage > 150000: score -= 8; issues.append("High mileage (" + str(mileage) + "km)")
    elif 0 < mileage < 80000: bonuses.append("Low mileage (" + str(mileage) + "km)")
    if price > 0 and purchase > 0:
        margin = (price - purchase) / price * 100
        if margin < 5:    score -= 10; issues.append("Low margin (" + str(round(margin)) + "%)")
        elif margin > 20: bonuses.append("Strong margin (" + str(round(margin)) + "%)")
    missing = [f for f in ["VIN","Engine","Color","Transmission","Body","Notes"] if not vehicle.get(f)]
    if len(missing) >= 4: score -= 15; issues.append("Missing fields: " + ", ".join(missing))
    elif len(missing) >= 2: score -= 5; issues.append("Missing: " + ", ".join(missing))
    else: bonuses.append("Complete listing data")
    if not vehicle.get("Key_Tag_ID"): score -= 5; issues.append("No key tag assigned")
    popular = ["Toyota","Mazda","Hyundai","Honda","Ford","Mitsubishi","Subaru","Kia","Volkswagen","BMW"]
    if vehicle.get("Make","").title() in popular:
        bonuses.append("Popular brand (" + vehicle.get("Make","").title() + ")")
    score = max(0, min(100, score))
    if score >= 80:   grade, color = "A", "#22c55e"
    elif score >= 60: grade, color = "B", "#84cc16"
    elif score >= 40: grade, color = "C", "#f59e0b"
    else:             grade, color = "D", "#ef4444"
    return {"score": score, "grade": grade, "color": color, "issues": issues, "bonuses": bonuses}

def gen_finance_options(vehicle, deposit=0, term_years=5):
    """Weekly/fortnightly/monthly repayments at 3 rates."""
    price_raw = re.sub(r"[^\d.]", "", str(vehicle.get("Price","") or ""))
    try: price = float(price_raw)
    except: return None
    principal = max(0, price - deposit)
    results = {}
    for rate_name, annual_rate in [("Low (6.9%)", 0.069), ("Avg (9.9%)", 0.099), ("High (13.9%)", 0.139)]:
        wr = annual_rate / 52
        nw = term_years * 52
        weekly = principal * wr / (1 - (1 + wr)**(-nw)) if wr > 0 else principal / nw
        results[rate_name] = {
            "weekly": round(weekly, 2), "fortnightly": round(weekly * 2, 2),
            "monthly": round(weekly * 52 / 12, 2), "total": round(weekly * nw, 2),
            "interest": round(weekly * nw - principal, 2),
        }
    return {"price": price, "deposit": deposit, "principal": principal,
            "term_years": term_years, "rates": results}

def gen_listing_full(vehicle, platform="carsales", tone="standard"):
    """Generate platform-specific listing copy."""
    if not openai_client:
        return "Add OPENAI_API_KEY to enable AI listings."
    ctx = (str(vehicle.get("Year","")) + " " + str(vehicle.get("Make","")) + " " +
           str(vehicle.get("Model","")) + " " + str(vehicle.get("Trim","")) +
           " — " + str(vehicle.get("Engine","")) + " — " +
           str(vehicle.get("Color","")) + " — " + str(vehicle.get("Mileage","")) +
           " — " + str(vehicle.get("Price","")))
    tone_guides = {
        "premium":     "Luxury, sophisticated tone. Focus on exclusivity and quality.",
        "budget":      "Value-focused. Emphasise affordability and practicality.",
        "family":      "Warm, safe, reliable. Focus on space, safety features, family suitability.",
        "performance": "Exciting, dynamic. Focus on power, handling, driving experience.",
        "standard":    "Professional, balanced. Facts-forward with a confident close.",
    }
    platform_guides = {
        "carsales":  "Carsales.com.au listing: structured, professional, 150-200 words. Include specs.",
        "facebook":  "Facebook Marketplace: casual, friendly, 80-120 words. Conversational.",
        "gumtree":   "Gumtree: brief, no-nonsense, 60-80 words. Price highlighted.",
        "instagram": "Instagram caption: punchy, emoji-friendly, 40-60 words + 5 relevant hashtags.",
    }
    prompt = ("Write a vehicle listing for: " + ctx + "\n" +
              "Notes: " + repr(str(vehicle.get("Notes",""))[:300]) + "\n" +
              "Platform: " + platform_guides.get(platform, platform_guides["carsales"]) + "\n" +
              "Tone: " + tone_guides.get(tone, tone_guides["standard"]) + "\n" +
              "Return ONLY the listing text, no intro or explanation.")
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role":"user","content":prompt}],
            max_tokens=400, temperature=0.75)
        return resp.choices[0].message.content.strip()
    except Exception as e:
        return "Error: " + str(e)

def gen_salesperson_brief(vehicle):
    """Full salesperson cheat sheet."""
    if not openai_client:
        return {"pitch":"Add OPENAI_API_KEY","objections":[],"talking_points":[],"why_buy_today":"","competitor_angle":""}
    ctx = (str(vehicle.get("Year","")) + " " + str(vehicle.get("Make","")) + " " +
           str(vehicle.get("Model","")) + " " + str(vehicle.get("Trim","")) +
           " — " + str(vehicle.get("Engine","")) + " — Mileage: " +
           str(vehicle.get("Mileage","")) + " — Price: " + str(vehicle.get("Price","")))
    prompt = ("You are an expert car sales trainer. For this vehicle: " + ctx + "\n" +
              "Return ONLY valid JSON with keys: " +
              "pitch (2 sentences), " +
              "talking_points (array of 5 short strings), " +
              "objections (array of 4 objects each with q and a keys), " +
              "why_buy_today (1 urgency sentence), " +
              "competitor_angle (1 sentence vs competitors)")
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role":"user","content":prompt}],
            max_tokens=600, temperature=0.7)
        raw = resp.choices[0].message.content.strip().replace("```json","").replace("```","")
        return json.loads(raw)
    except Exception as e:
        return {"pitch":str(e),"objections":[],"talking_points":[],"why_buy_today":"","competitor_angle":""}

def calc_profit(vehicle):
    try:
        price    = float(re.sub(r"[^\d.]", "", str(vehicle.get("Price","") or "")) or 0)
        purchase = float(re.sub(r"[^\d.]", "", str(vehicle.get("Purchase_Price","") or "")) or 0)
        recon    = float(re.sub(r"[^\d.]", "", str(vehicle.get("Recon_Cost","") or "")) or 0)
        if price > 0 and purchase > 0:
            profit = price - purchase - recon
            margin = round(profit / price * 100, 1)
            return {"profit": round(profit), "margin_pct": margin, "has_data": True}
    except Exception:
        pass
    return {"profit": None, "margin_pct": None, "has_data": False}

def maybe_log_price(internal_id, new_price, user):
    if not new_price or not internal_id: return
    history = _price_history[internal_id.upper()]
    last = history[-1]["price"] if history else None
    if new_price != last:
        history.append({
            "price": new_price,
            "date": datetime.now(timezone.utc).strftime("%b %d %Y"),
            "user": user,
        })
        # Notify buyers who saved this vehicle if price dropped
        if last is not None:
            try:
                old_p = float(re.sub(r"[^\d.]","",str(last)) or 0)
                new_p = float(re.sub(r"[^\d.]","",str(new_price)) or 0)
                if new_p < old_p and new_p > 0:
                    _notify_buyers_price_drop(internal_id.upper(), old_p, new_p)
            except Exception:
                pass

# ── NHTSA ──────────────────────────────────────────────────────────────────────

NHTSA_DISPLAY = {
    "Identity":   ["Make","Model","ModelYear","BodyClass","VehicleType","Trim","Series","Manufacturer","PlantCity","PlantCountry"],
    "Powertrain": ["DisplacementL","EngineCylinders","EngineHP","FuelTypePrimary","TransmissionStyle","DriveType","GVWR"],
    "Dimensions": ["Doors","WheelBaseShort","TrackFront"],
    "Safety":     ["ABS","ESC","TractionControl","ForwardCollisionWarning","LaneDepartureWarning",
                   "BackupCamera","BlindSpotMon","AdaptiveCruiseControl","KeylessIgnition",
                   "DaytimeRunningLight","AirBagLocFront","AirBagLocSide"],
}
NHTSA_LABELS = {
    "Make":"Make","Model":"Model","ModelYear":"Year","BodyClass":"Body Style",
    "VehicleType":"Vehicle Type","Trim":"Trim","Series":"Series",
    "Manufacturer":"Manufacturer","PlantCity":"Assembly City","PlantCountry":"Assembly Country",
    "DisplacementL":"Displacement (L)","EngineCylinders":"Cylinders","EngineHP":"Horsepower",
    "FuelTypePrimary":"Fuel Type","TransmissionStyle":"Transmission","DriveType":"Drive Type",
    "GVWR":"GVWR","Doors":"Doors","WheelBaseShort":"Wheelbase (in)","TrackFront":"Front Track (in)",
    "ABS":"ABS","ESC":"Stability Control","TractionControl":"Traction Control",
    "ForwardCollisionWarning":"Forward Collision Warning","LaneDepartureWarning":"Lane Departure Warning",
    "BackupCamera":"Backup Camera","BlindSpotMon":"Blind Spot Monitor",
    "AdaptiveCruiseControl":"Adaptive Cruise Control","KeylessIgnition":"Keyless Ignition",
    "DaytimeRunningLight":"Daytime Running Lights","AirBagLocFront":"Airbags (Front)","AirBagLocSide":"Airbags (Side)",
}
SECTION_ICONS = {"Identity":"🚗","Powertrain":"⚙️","Dimensions":"📐","Safety":"🛡️"}

def decode_vin(vin):
    vin = vin.strip().upper()
    if len(vin) < 11: return {"ok": False, "error": "VIN too short"}
    try:
        r = requests.get(NHTSA_VIN_URL.format(vin=vin), timeout=10)
        if r.status_code != 200: return {"ok": False, "error": f"NHTSA error {r.status_code}"}
        raw = r.json().get("Results", [{}])[0]
        if not raw.get("Make"): return {"ok": False, "error": "VIN not recognized"}
        def clean(v): return "" if v in ("Not Applicable","0","","N/A") else str(v).strip()
        sections = {}
        for section, keys in NHTSA_DISPLAY.items():
            items = {NHTSA_LABELS[k]: clean(raw.get(k,"")) for k in keys if clean(raw.get(k,""))}
            if items: sections[f"{SECTION_ICONS[section]} {section}"] = items
        hp   = raw.get("EngineHP","").split(".")[0]
        cyls = raw.get("EngineCylinders","")
        disp = raw.get("DisplacementL","")
        parts = [p for p in [
            f"{disp}L" if disp and disp!="0" else "",
            f"V{cyls}" if cyls and int(cyls or 0)>4 else (f"{cyls}-cyl" if cyls and cyls!="0" else ""),
            f"{hp}hp"  if hp and hp!="0" else ""
        ] if p]
        return {
            "ok": True,
            "make": raw.get("Make","").title(), "model": raw.get("Model","").title(),
            "year": raw.get("ModelYear",""), "trim": clean(raw.get("Trim","")),
            "engine": " ".join(parts), "body": clean(raw.get("BodyClass","")),
            "drive": clean(raw.get("DriveType","")), "fuel": clean(raw.get("FuelTypePrimary","")),
            "transmission": clean(raw.get("TransmissionStyle","")),
            "doors": clean(raw.get("Doors","")), "sections": sections,
        }
    except requests.Timeout:
        return {"ok": False, "error": "NHTSA API timed out"}
    except Exception as e:
        return {"ok": False, "error": str(e)}

def get_recalls(make, model, year):
    if not all([make, model, year]): return []
    try:
        url = NHTSA_RECALL_URL.format(
            make=requests.utils.quote(str(make)),
            model=requests.utils.quote(str(model)),
            year=str(year))
        r = requests.get(url, timeout=8)
        if r.status_code != 200: return []
        return [{"campaign": x.get("NHTSACampaignNumber",""), "date": x.get("ReportReceivedDate",""),
                 "component": x.get("Component",""), "summary": x.get("Summary",""),
                 "consequence": x.get("Consequence",""), "remedy": x.get("Remedy","")}
                for x in r.json().get("results", [])]
    except Exception:
        return []

# ── OpenAI ─────────────────────────────────────────────────────────────────────

def gen_pitch(vehicle):
    if not openai_client: return "Add OPENAI_API_KEY to .env to enable AI pitches."
    ctx = f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')} {vehicle.get('Trim','')}".strip()
    if vehicle.get("Engine"): ctx += f" · {vehicle['Engine']}"
    if vehicle.get("Price"):  ctx += f" · {vehicle['Price']}"
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role":"user","content":
                f"You are a top car salesperson. Write a punchy 3-sentence pitch for: {ctx}.\n"
                f"Notes: \"{vehicle.get('Notes','No notes.')}\"\n"
                f"Be specific, create urgency, end with a reason to buy TODAY. No clichés."}],
            max_tokens=200, temperature=0.75)
        return resp.choices[0].message.content.strip()
    except Exception as e:
        return f"Error: {e}"

def gen_objections(vehicle):
    if not openai_client: return []
    ctx = f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')} at {vehicle.get('Price','')}"
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role":"user","content":
                f"For a {ctx}, give 3 one-liner comebacks for: (1) too expensive, "
                f"(2) need to think about it, (3) cheaper elsewhere. "
                f"Return ONLY a valid JSON array of 3 strings."}],
            max_tokens=200, temperature=0.7)
        raw = resp.choices[0].message.content.strip().replace("```json","").replace("```","")
        return json.loads(raw)
    except Exception:
        return []

def gen_listing_draft(vehicle):
    if not openai_client: return "Add OPENAI_API_KEY to enable listing drafts."
    ctx = (f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')} "
           f"{vehicle.get('Trim','')} — {vehicle.get('Engine','')} — {vehicle.get('Color','')} — "
           f"{vehicle.get('Mileage','')} — {vehicle.get('Price','')}")
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role":"user","content":
                f"Write a professional vehicle listing for: {ctx}\n"
                f"Notes: \"{vehicle.get('Notes','')}\"\n"
                f"Format: Title line, 3 punchy bullet features, 2-sentence call to action. "
                f"Suitable for Trade-Me/AutoTrader. No markdown headers."}],
            max_tokens=350, temperature=0.7)
        return resp.choices[0].message.content.strip()
    except Exception as e:
        return f"Error: {e}"

# ── Twilio SMS ─────────────────────────────────────────────────────────────────

def send_sms(to, body):
    if not all([TWILIO_SID, TWILIO_TOKEN_ENV, TWILIO_FROM, to]):
        return False, "Twilio not configured"
    try:
        r = requests.post(
            f"https://api.twilio.com/2010-04-01/Accounts/{TWILIO_SID}/Messages.json",
            auth=(TWILIO_SID, TWILIO_TOKEN_ENV),
            data={"From": TWILIO_FROM, "To": to, "Body": body}, timeout=10)
        return r.status_code in (200,201), r.text[:200]
    except Exception as e:
        return False, str(e)

# ── Auth ───────────────────────────────────────────────────────────────────────

def current_user():
    return session.get("username", "staff")

def is_auth():
    if session.get("authenticated"): return True
    if URL_SECRET_TOKEN and request.args.get("token") == URL_SECRET_TOKEN:
        session["authenticated"] = True; return True
    return False

def login_required(f):
    @wraps(f)
    def d(*a, **kw):
        if not is_auth():
            if request.method == "POST" or request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"ok": False, "message": "Session expired — please log in again"}), 401
            return redirect(url_for("login", next=request.url))
        return f(*a, **kw)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a, **kw):
        if not session.get("is_admin"):
            if request.method == "POST" or request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"ok": False, "message": "Admin access required"}), 401
            return redirect(url_for("admin_login"))
        return f(*a, **kw)
    return d

# ══════════════════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    if is_auth(): return redirect(url_for("dashboard"))
    return redirect("/home")

# ── Auth ───────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET","POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username","").strip().lower()
        password = request.form.get("password","")
        acct = _staff_accounts.get(username)
        if acct and check_password_hash(acct["password"], password):
            session.update({"authenticated":True, "username":username,
                            "display_name":acct["name"], "role":acct["role"],
                            "is_admin": acct["role"]=="admin"})
            next_url = request.args.get("next","")
            if not next_url.startswith("/"):
                next_url = url_for("dashboard")
            return redirect(next_url)
        # Legacy single-password fallback (also hashed check)
        if check_password_hash(_staff_accounts.get("staff",{}).get("password",""), password):
            session.update({"authenticated":True, "username":"staff",
                            "display_name":"Staff", "role":"staff", "is_admin":False})
            return redirect(url_for("dashboard"))
        error = "Invalid username or password."
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("login"))

# ── Dashboard ──────────────────────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    vehicles = get_all_vehicles(include_archived=False)
    stats = {
        "total": len(vehicles),
        "available": sum(1 for v in vehicles if v.get("Status")=="Available"),
        "sold": sum(1 for v in vehicles if v.get("Status")=="Sold"),
        "service": sum(1 for v in vehicles if v.get("Status")=="In Service"),
        "aging_30": sum(1 for v in vehicles if (v.get("_days_on_lot") or 0) >= 30),
    }
    upcoming_drives = [td for td in _test_drives if td.get("status")=="scheduled"][:5]
    target     = _settings.get("monthly_target", 15)
    sold_month = sold_this_month(vehicles)
    followups  = due_followups()
    overdue    = overdue_keys()
    return render_template("dashboard.html", vehicles=vehicles, stats=stats,
                           dealership=DEALERSHIP_NAME, upcoming_drives=upcoming_drives,
                           user=session.get("display_name","Staff"),
                           overdue_keys=overdue, followups=followups,
                           sold_month=sold_month, monthly_target=target,
                           settings=_settings)

# ── Vehicle scan ───────────────────────────────────────────────────────────────

@app.route("/scan/<internal_id>")
@login_required
def scan(internal_id):
    vehicle = get_vehicle(internal_id.upper())
    if not vehicle: abort(404)
    pitch      = gen_pitch(vehicle)
    objections = gen_objections(vehicle)
    recalls    = get_recalls(vehicle.get("Make",""), vehicle.get("Model",""), str(vehicle.get("Year","")))
    profit     = calc_profit(vehicle)
    dot = None
    if vehicle.get("Date_Added"):
        try:
            added = datetime.fromisoformat(vehicle["Date_Added"].replace("Z",""))
            dot = (datetime.now() - added).days
        except Exception: pass
    enquiries_for_v = [e for e in _enquiries if e["internal_id"]==internal_id.upper()]
    drives_for_v    = [td for td in _test_drives if td["internal_id"]==internal_id.upper()]
    service_for_v   = [s for s in _service_log if s["internal_id"]==internal_id.upper()]
    price_hist      = _price_history.get(internal_id.upper(), [])
    tag = vehicle.get("Key_Tag_ID","")
    key_status = _key_checkouts.get(tag) if tag else None
    log_activity("scan", internal_id.upper(),
                 f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')}", current_user())
    # Build the staff-personalised share URL — includes ?ref= so customer sees their number
    _share_base = request.host_url.rstrip("/")
    _ref = session.get("username", "")
    share_url_staff = f"{_share_base}/share/{internal_id.upper()}?ref={_ref}" if _ref else f"{_share_base}/share/{internal_id.upper()}"
    return render_template("vehicle.html",
        vehicle=vehicle, pitch=pitch, objections=objections, recalls=recalls,
        internal_id=internal_id.upper(), dealership=DEALERSHIP_NAME,
        profit=profit, days_on_lot=dot, dot_flag=dot_flag(dot),
        enquiries=enquiries_for_v, test_drives=drives_for_v,
        service_log=service_for_v, price_history=price_hist,
        key_status=key_s,
        share_url_staff=share_url_stafftatus, is_admin=session.get("is_admin", False),
        statuses=VEHICLE_STATUSES)

@app.route("/share/<internal_id>")
def vehicle_share(internal_id):
    """Public shareable spec sheet — no login needed.
    Pass ?ref=<username> to show that staff member's personal contact number.
    Admin can update any staff phone at any time (Admin → Staff) so the
    customer always reaches the right person even if they changed their number.
    """
    vehicle = get_vehicle(internal_id.upper())
    if not vehicle: abort(404)
    if vehicle.get("Status") == "Archived": abort(404)

    # Resolve staff member from ?ref= query param
    ref_username = request.args.get("ref", "").strip().lower()
    staff_contact = None
    if ref_username and ref_username in _staff_accounts:
        acct = _staff_accounts[ref_username]
        phone = acct.get("phone", "").strip()
        email = acct.get("email", "").strip()
        # Only expose name + phone + email (never password hash or role internals)
        staff_contact = {
            "name":     acct.get("name", ref_username.title()),
            "username": ref_username,
            "phone":    phone,
            "email":    email,
            "has_phone": bool(phone),
        }

    log_activity("share_view", internal_id.upper(),
                 f"Shared by {ref_username or 'unknown'}", ref_username or "public")
    return render_template("vehicle_share.html",
        vehicle=vehicle,
        internal_id=internal_id.upper(),
        dealership=DEALERSHIP_NAME,
        staff_contact=staff_contact,
        ref_username=ref_username,
        dealership_phone=DEALERSHIP_PHONE or MANAGER_PHONE,
        dealership_address=DEALERSHIP_ADDRESS)

@app.route("/customers")
@login_required
def customers_page():
    """CRM-style view of all unique customers across all enquiries."""
    # Group enquiries by customer name+phone
    from collections import defaultdict
    cust_map = defaultdict(lambda: {"enquiries": [], "vehicles": set()})
    for e in _enquiries:
        key = (e.get("name","").lower().strip(), e.get("phone","").strip())
        cust_map[key]["name"]   = e.get("name","")
        cust_map[key]["phone"]  = e.get("phone","")
        cust_map[key]["email"]  = e.get("email","")
        cust_map[key]["enquiries"].append(e)
        if e.get("internal_id"):
            cust_map[key]["vehicles"].add(e["internal_id"])
    customers = []
    for key, data in cust_map.items():
        enqs = sorted(data["enquiries"], key=lambda x: x.get("timestamp",""), reverse=True)
        latest = enqs[0]
        pending_fu = any(
            e.get("follow_up_date") and e.get("follow_up_status") == "pending"
            for e in enqs
        )
        overdue_fu = any(
            e.get("follow_up_date","") <= datetime.now().strftime("%Y-%m-%d")
            and e.get("follow_up_status") == "pending"
            and e.get("follow_up_date","") != ""
            for e in enqs
        )
        customers.append({
            "name":        data.get("name",""),
            "phone":       data.get("phone",""),
            "email":       data.get("email",""),
            "vehicles":    sorted(data["vehicles"]),
            "enquiry_count": len(enqs),
            "latest_note": latest.get("notes",""),
            "last_contact": latest.get("timestamp",""),
            "follow_up_date": latest.get("follow_up_date",""),
            "follow_up_status": latest.get("follow_up_status","pending"),
            "overdue":     overdue_fu,
            "pending_fu":  pending_fu,
            "enq_ids":     [e["id"] for e in enqs],
        })
    customers.sort(key=lambda x: (not x["overdue"], x.get("last_contact","") ), reverse=False)
    customers.sort(key=lambda x: x["overdue"], reverse=True)
    return render_template("customers.html", customers=customers,
                           dealership=DEALERSHIP_NAME,
                           followup_count=sum(1 for c in customers if c["overdue"]))

@app.route("/keyboard")
@login_required
def keyboard_page():
    """Live key board — wall display. No login if URL token present."""
    # Use same auth as rest of app (is_auth checks session["authenticated"])
    if not is_auth():
        return redirect(url_for("login", next=request.url))
    # Get all vehicles for tag→vehicle lookup
    all_v = get_all_vehicles(include_archived=False)
    tag_to_vehicle = {v["Key_Tag_ID"]: v for v in all_v if v.get("Key_Tag_ID")}
    overdue = overdue_keys()
    overdue_tags = {o["tag"] for o in overdue}
    checkouts = []
    for tag, co in _key_checkouts.items():
        vehicle = tag_to_vehicle.get(tag, {})
        try:
            ts = datetime.fromisoformat(co["timestamp"])
            if ts.tzinfo is None: ts = ts.replace(tzinfo=timezone.utc)
            mins = int((datetime.now(timezone.utc) - ts).total_seconds() / 60)
        except Exception:
            mins = 0
        checkouts.append({
            **co,
            "vehicle_name": f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')}".strip(),
            "minutes_out": mins,
            "overdue": tag in overdue_tags,
        })
    checkouts.sort(key=lambda x: x["minutes_out"], reverse=True)
    # Available keys = vehicles with key tags not currently checked out
    available = [v for v in all_v if v.get("Key_Tag_ID") and v["Key_Tag_ID"] not in _key_checkouts]
    return render_template("keyboard.html",
                           checkouts=checkouts, available=available,
                           overdue_count=len(overdue),
                           dealership=DEALERSHIP_NAME,
                           threshold=_settings.get("key_overdue_hours", 2))

@app.route("/settings", methods=["GET","POST"])
@admin_required
def settings_page():
    if request.method == "POST":
        d = request.get_json(silent=True) or {}
        if "monthly_target" in d:
            try:
                v = int(d["monthly_target"])
                _settings["monthly_target"] = max(1, min(500, v))
            except (ValueError, TypeError): pass
        if "key_overdue_hours" in d:
            try:
                v = int(d["key_overdue_hours"])
                _settings["key_overdue_hours"] = max(1, min(72, v))
            except (ValueError, TypeError): pass
        if "dealership_name"   in d: _settings["dealership_name"]   = d["dealership_name"].strip()
        return jsonify({"ok": True, "settings": _settings})
    return render_template("settings.html", settings=_settings, dealership=DEALERSHIP_NAME)

@app.route("/api/followup-done", methods=["POST"])
@login_required
def api_followup_done():
    d = request.get_json(silent=True) or {}
    enq_id = d.get("id","")
    for e in _enquiries:
        if e.get("id") == enq_id:
            e["follow_up_status"] = "done"
            log_activity("followup_done", e.get("internal_id",""), f"Follow-up marked done for {e.get('name','')}", current_user())
            return jsonify({"ok": True})
    return jsonify({"ok": False, "message": "Enquiry not found"})

@app.route("/api/set-followup", methods=["POST"])
@login_required
def api_set_followup():
    d = request.get_json(silent=True) or {}
    enq_id = d.get("id","")
    date   = d.get("date","")
    for e in _enquiries:
        if e.get("id") == enq_id:
            e["follow_up_date"]   = date
            e["follow_up_status"] = "pending"
            return jsonify({"ok": True})
    return jsonify({"ok": False, "message": "Enquiry not found"})

@app.route("/api/qr-scan/<vid>")
def api_qr_scan_image(vid):
    """QR PNG for a vehicle — server-side generation. No login needed (URL not secret)."""
    import io as _io
    vid  = vid.upper().strip()
    base = request.host_url.rstrip("/")
    url  = f"{base}/scan/{vid}"

    if HAS_QRCODE:
        try:
            qr = qrcode.QRCode(
                version=None,
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=10,
                border=4,
            )
            qr.add_data(url)
            qr.make(fit=True)
            img = qr.make_image(image_factory=_PyPNG)
            buf = _io.BytesIO()
            img.save(buf)
            buf.seek(0)
            from flask import send_file
            resp = send_file(buf, mimetype="image/png", download_name=f"QR_{vid}.png")
            resp.headers["Cache-Control"] = "public, max-age=3600"
            return resp
        except Exception:
            pass

    # Fallback: redirect to a reliable QR API if qrcode lib not available
    encoded = requests.utils.quote(url, safe="")
    fallback = f"https://api.qrserver.com/v1/create-qr-code/?size=300x300&ecc=M&data={encoded}"
    return redirect(fallback)

@app.route("/scanner")
@login_required
def scanner_page():
    return render_template("scanner.html", dealership=DEALERSHIP_NAME)

@app.route("/api/vehicle/<vid>")
@login_required
def api_vehicle_lookup(vid):
    """Quick vehicle lookup for the scanner page."""
    vid = vid.upper().strip()
    url = (f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{AIRTABLE_TABLE}"
           f"?filterByFormula={{Internal_ID}}='{vid}'&maxRecords=1")
    try:
        r = requests.get(url, headers=at_headers(), timeout=8)
        if r.status_code == 200:
            records = r.json().get("records", [])
            if records:
                f = records[0].get("fields", {})
                f["_id"] = records[0]["id"]
                # Strip private fields
                safe = {k: v for k, v in f.items() if not k.startswith("_")}
                return jsonify({"ok": True, "vehicle": safe})
    except Exception as ex:
        return jsonify({"ok": False, "message": str(ex)}), 500
    return jsonify({"ok": False, "message": "Not found"}), 404

# /vin-lookup page removed by request

@app.route("/compare")
@login_required
def compare_page():
    ids = request.args.getlist("ids")
    vehicles = [v for vid in ids[:3] if (v := get_vehicle(vid.upper()))]
    all_v = get_all_vehicles()
    return render_template("compare.html", vehicles=vehicles, all_vehicles=all_v, dealership=DEALERSHIP_NAME)

# ── AJAX core ──────────────────────────────────────────────────────────────────

@app.route("/api/update-status", methods=["POST"])
@login_required
def api_update_status():
    d = request.get_json()
    internal_id = d.get("internal_id","").strip().upper()
    status_val  = clean_status(d.get("status",""))
    if not internal_id:
        return jsonify({"ok": False, "message": "Missing vehicle ID"})
    if status_val not in VEHICLE_STATUSES:
        return jsonify({"ok": False, "message": f"Invalid status: {status_val}"})

    # ── Sold: save all buyer capture fields ──────────────────────────────────
    if status_val == "Sold":
        fields = {
            "Status":        "Sold",
            "Sold_Date":     datetime.now().strftime("%Y-%m-%d"),
            "Sold_By":       current_user(),
        }
        # Map every buyer field from the payload
        field_map = {
            "buyer_name":      "Buyer_Name",
            "buyer_phone":     "Buyer_Phone",
            "buyer_dob":       "Buyer_DOB",
            "buyer_licence":   "Buyer_Licence",
            "buyer_card_last4":"Buyer_Card_Last4",
            "buyer_card_expiry":"Buyer_Card_Expiry",
            "sale_price":      "Sale_Price",
            "sale_notes":      "Sale_Notes",
        }
        for payload_key, airtable_key in field_map.items():
            val = d.get(payload_key, "").strip() if isinstance(d.get(payload_key), str) else ""
            if val:
                fields[airtable_key] = val
        # Store live photo as base64 in a long-text field (Airtable attachment
        # upload requires separate pre-signed URL flow — store as data-URI for now)
        if d.get("buyer_photo"):
            fields["Buyer_Photo_Data"] = d["buyer_photo"][:65000]  # Airtable long-text limit safe

        ok, msg = patch_vehicle(internal_id, fields, user=current_user())
        if ok:
            vehicle = get_vehicle(internal_id)
            log_activity("sold", internal_id,
                         f"Sold by {current_user()} to {fields.get('Buyer_Name','Unknown')}", current_user())
            sms = (f"🏆 SOLD: {vehicle.get('Year','')} {vehicle.get('Make','')} "
                   f"{vehicle.get('Model','')} — {vehicle.get('Price','')} "
                   f"by {current_user()} [{internal_id}]")
            send_sms(MANAGER_PHONE, sms)
        return jsonify({"ok": ok, "message": msg if not ok else "Vehicle marked as Sold!"})
    # ─────────────────────────────────────────────────────────────────────────

    ok, msg = patch_vehicle(internal_id, {"Status": status_val}, user=current_user(), audit=False)
    if ok:
        log_activity("status_change", internal_id, f"→ {status_val}", current_user())
        return jsonify({"ok": True, "message": f"Status updated to {status_val}"})
    return jsonify({"ok": False, "message": f"Airtable error: {msg}"})

@app.route("/api/regenerate-pitch", methods=["POST"])
@login_required
def api_regenerate_pitch():
    vehicle = get_vehicle(request.get_json().get("internal_id",""))
    if not vehicle: return jsonify({"ok":False,"error":"Not found"}), 404
    return jsonify({"ok":True,"pitch":gen_pitch(vehicle),"objections":gen_objections(vehicle)})

@app.route("/api/decode-vin", methods=["POST"])
@login_required
def api_decode_vin():
    vin = request.get_json().get("vin","").strip().upper()
    result = decode_vin(vin)
    if result.get("ok"):
        log_activity("vin_decode","",f"VIN {vin} → {result.get('year','')} {result.get('make','')} {result.get('model','')}")
    return jsonify(result)

@app.route("/api/check-recalls", methods=["POST"])
@login_required
def api_check_recalls():
    d = request.get_json()
    recalls = get_recalls(d.get("make",""), d.get("model",""), d.get("year",""))
    return jsonify({"ok":True,"recalls":recalls,"count":len(recalls)})

@app.route("/api/activity-log")
@login_required
def api_activity_log():
    return jsonify({"ok":True,"log":_activity_log[:int(request.args.get("limit",50))]})

# ── Staff notes ────────────────────────────────────────────────────────────────

@app.route("/api/add-note", methods=["POST"])
@login_required
def api_add_note():
    d = request.get_json()
    vid  = d.get("internal_id","").upper()
    note = d.get("note","").strip()
    if not note: return jsonify({"ok":False,"message":"Note cannot be empty"})
    vehicle = get_vehicle(vid)
    if not vehicle: return jsonify({"ok":False,"message":"Vehicle not found"})
    ts   = datetime.now(timezone.utc).strftime("%b %d %H:%M")
    user = session.get("display_name","Staff")
    new_notes = (vehicle.get("Notes","") + f"\n\n[{ts} — {user}] {note}").strip()
    ok, msg = patch_vehicle(vid, {"Notes": new_notes}, user=current_user(), audit=False)
    if ok: log_activity("note_added", vid, note[:60], current_user())
    return jsonify({"ok":ok,"message":msg,"notes":new_notes})

# ── Customer enquiries ─────────────────────────────────────────────────────────

@app.route("/api/add-enquiry", methods=["POST"])
@login_required
def api_add_enquiry():
    d = request.get_json()
    if not d.get("name","").strip():
        return jsonify({"ok":False,"message":"Customer name required"})
    enquiry = {
        "id": str(uuid.uuid4())[:8],
        "internal_id": d.get("internal_id","").upper(),
        "name": d.get("name","").strip(),
        "phone": d.get("phone","").strip(),
        "email": d.get("email","").strip(),
        "notes": d.get("notes","").strip(),
        "follow_up_date": d.get("follow_up_date","").strip(),
        "follow_up_status": "pending",
        "timestamp": datetime.now(timezone.utc).strftime("%b %d %H:%M"),
        "timestamp_iso": datetime.now(timezone.utc).isoformat(),
        "staff": session.get("display_name","Staff"),
    }
    _enquiries.insert(0, enquiry)
    log_activity("enquiry", enquiry["internal_id"], f"{enquiry['name']} — {enquiry['phone']}", current_user())
    return jsonify({"ok":True,"enquiry":enquiry})

@app.route("/api/enquiries/<internal_id>")
@login_required
def api_get_enquiries(internal_id):
    return jsonify({"ok":True,"enquiries":[e for e in _enquiries if e["internal_id"]==internal_id.upper()]})

@app.route("/api/all-enquiries")
@admin_required
def api_all_enquiries():
    return jsonify({"ok":True,"enquiries":_enquiries[:200]})

# ── Test drive bookings ────────────────────────────────────────────────────────

@app.route("/api/book-test-drive", methods=["POST"])
@login_required
def api_book_test_drive():
    d = request.get_json()
    vid = d.get("internal_id","").upper()
    dt  = d.get("datetime","").strip()
    customer = d.get("customer_name","").strip()
    if not all([vid, dt, customer]):
        return jsonify({"ok":False,"message":"Vehicle, date/time, and customer name required"})
    conflict = any(td["internal_id"]==vid and td["datetime"]==dt and td["status"]=="scheduled"
                   for td in _test_drives)
    if conflict:
        return jsonify({"ok":False,"message":"Another test drive already booked at that time"})
    booking = {
        "id": str(uuid.uuid4())[:8], "internal_id": vid, "datetime": dt,
        "customer_name": customer, "phone": d.get("phone","").strip(),
        "notes": d.get("notes","").strip(), "status": "scheduled",
        "staff": session.get("display_name","Staff"),
        "created": datetime.now(timezone.utc).isoformat(),
    }
    _test_drives.insert(0, booking)
    log_activity("test_drive_booked", vid, f"{customer} @ {dt}", current_user())
    return jsonify({"ok":True,"booking":booking})

@app.route("/api/test-drives/<internal_id>")
@login_required
def api_get_test_drives(internal_id):
    return jsonify({"ok":True,"bookings":[td for td in _test_drives if td["internal_id"]==internal_id.upper()]})

@app.route("/api/test-drives/update", methods=["POST"])
@login_required
def api_update_test_drive():
    d = request.get_json()
    for td in _test_drives:
        if td["id"] == d.get("id"):
            td["status"] = d.get("status","scheduled")
            return jsonify({"ok":True})
    return jsonify({"ok":False,"message":"Booking not found"})

@app.route("/api/all-test-drives")
@login_required
def api_all_test_drives():
    return jsonify({"ok":True,"bookings":_test_drives[:100]})

# ── Service log ────────────────────────────────────────────────────────────────

@app.route("/api/add-service", methods=["POST"])
@login_required
def api_add_service():
    d = request.get_json()
    vid = d.get("internal_id","").upper()
    if not d.get("description","").strip():
        return jsonify({"ok":False,"message":"Description required"})
    entry = {
        "id": str(uuid.uuid4())[:8], "internal_id": vid,
        "type": d.get("type","service"),
        "description": d.get("description","").strip(),
        "cost": d.get("cost","").strip(),
        "date": d.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        "technician": d.get("technician","").strip(),
        "staff": session.get("display_name","Staff"),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    _service_log.insert(0, entry)
    log_activity("service_entry", vid, f"{entry['type']}: {entry['description'][:40]}", current_user())
    return jsonify({"ok":True,"entry":entry})

@app.route("/api/service-log/<internal_id>")
@login_required
def api_get_service_log(internal_id):
    return jsonify({"ok":True,"entries":[s for s in _service_log if s["internal_id"]==internal_id.upper()]})

# ── Key tag checkout ───────────────────────────────────────────────────────────

@app.route("/api/key-checkout", methods=["POST"])
@login_required
def api_key_checkout():
    d = request.get_json()
    tag    = d.get("tag_id","").strip()
    vid    = d.get("internal_id","").upper()
    action = d.get("action","checkout")
    staff  = session.get("display_name","Staff")
    if action == "checkout":
        if tag in _key_checkouts:
            co = _key_checkouts[tag]
            return jsonify({"ok":False,"message":f"Key already out with {co['staff']} since {co['time']}"})
        _key_checkouts[tag] = {
            "tag": tag, "vehicle": vid, "staff": staff,
            "username": current_user(),
            "time": datetime.now(timezone.utc).strftime("%b %d %H:%M"),
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        log_activity("key_checkout", vid, f"Tag {tag} — {staff}", current_user())
        return jsonify({"ok":True,"message":f"Key {tag} checked out to {staff}"})
    else:
        if tag not in _key_checkouts:
            return jsonify({"ok":False,"message":"Key not recorded as checked out"})
        _key_checkouts.pop(tag)
        log_activity("key_checkin", vid, f"Tag {tag} returned by {staff}", current_user())
        return jsonify({"ok":True,"message":f"Key {tag} returned"})

@app.route("/api/key-status/<tag_id>")
@login_required
def api_key_status(tag_id):
    status = _key_checkouts.get(tag_id.strip())
    return jsonify({"ok":True,"checked_out":bool(status),"info":status or {}})

@app.route("/api/all-checkouts")
@login_required
def api_all_checkouts():
    return jsonify({"ok":True,"checkouts":list(_key_checkouts.values())})

# ── Price history ──────────────────────────────────────────────────────────────

@app.route("/api/price-history/<internal_id>")
@login_required
def api_price_history(internal_id):
    return jsonify({"ok":True,"history":_price_history.get(internal_id.upper(),[])})

# ── Listing draft ──────────────────────────────────────────────────────────────

@app.route("/api/listing-draft", methods=["POST"])
@login_required
def api_listing_draft():
    vehicle = get_vehicle(request.get_json().get("internal_id",""))
    if not vehicle: return jsonify({"ok":False,"error":"Not found"})
    return jsonify({"ok":True,"draft":gen_listing_draft(vehicle)})

# ── CSV Export ─────────────────────────────────────────────────────────────────

@app.route("/admin/export-csv")
@admin_required
def admin_export_csv():
    all_v = get_all_vehicles(include_archived=True)
    fields = ["Internal_ID","Make","Model","Year","Color","VIN","Mileage","Price",
              "Trim","Engine","Status","Key_Tag_ID","Notes","Date_Added",
              "Purchase_Price","Recon_Cost"]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    for v in all_v:
        writer.writerow({f: v.get(f,"") for f in fields})
    resp = make_response(output.getvalue())
    resp.headers["Content-Type"] = "text/csv"
    resp.headers["Content-Disposition"] = (
        f"attachment; filename=keytrack-export-{datetime.now().strftime('%Y%m%d')}.csv")
    log_activity("csv_export","","Full inventory exported","admin")
    return resp

# ── Backup ─────────────────────────────────────────────────────────────────────

@app.route("/admin/backup")
@admin_required
def admin_backup():
    all_v = get_all_vehicles(include_archived=True)
    backup = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "dealership": DEALERSHIP_NAME,
        "vehicles": all_v,
        "enquiries": _enquiries,
        "test_drives": _test_drives,
        "service_log": _service_log,
        "activity_log": _activity_log[:200],
    }
    resp = make_response(json.dumps(backup, indent=2, default=str))
    resp.headers["Content-Type"] = "application/json"
    resp.headers["Content-Disposition"] = (
        f"attachment; filename=keytrack-backup-{datetime.now().strftime('%Y%m%d-%H%M')}.json")
    log_activity("backup","","Full JSON backup downloaded","admin")
    return resp

# ── Reports ────────────────────────────────────────────────────────────────────

@app.route("/reports")
@admin_required
def reports_page():
    all_v = get_all_vehicles(include_archived=True)
    sold  = [v for v in all_v if v.get("Status")=="Sold"]
    lots  = [v.get("_days_on_lot") for v in all_v if v.get("_days_on_lot") is not None and v.get("Status")=="Available"]
    avg_lot = round(sum(lots)/len(lots)) if lots else 0
    aging = {
        "90+ days": sum(1 for d in lots if d>=90),
        "60-90 days": sum(1 for d in lots if 60<=d<90),
        "30-60 days": sum(1 for d in lots if 30<=d<60),
        "Under 30": sum(1 for d in lots if d<30),
    }
    profits = [p["profit"] for v in all_v if (p := calc_profit(v))["has_data"] and p["profit"] is not None]
    total_profit = round(sum(profits)) if profits else 0
    avg_profit   = round(total_profit/len(profits)) if profits else 0
    make_breakdown = Counter(v.get("Make","Unknown") for v in all_v if v.get("Make")).most_common(10)
    return render_template("reports.html",
        all_vehicles=all_v, avg_lot=avg_lot, aging=aging,
        total_profit=total_profit, avg_profit=avg_profit,
        make_breakdown=make_breakdown,
        enquiries=_enquiries[:50], test_drives=_test_drives[:50],
        service_log=_service_log[:50],
        dealership=DEALERSHIP_NAME)

# ── Audit trail ────────────────────────────────────────────────────────────────

@app.route("/admin/audit")
@admin_required
def admin_audit():
    vid = request.args.get("vehicle","")
    trail = [e for e in _audit_trail if not vid or e["internal_id"]==vid.upper()]
    return render_template("audit.html", trail=trail[:500], filter_vehicle=vid, dealership=DEALERSHIP_NAME)

@app.route("/api/audit-log")
@admin_required
def api_audit_log():
    vid = request.args.get("vehicle","")
    data = [e for e in _audit_trail if not vid or e["internal_id"]==vid.upper()]
    return jsonify({"ok":True,"trail":data[:200]})

# ── Staff management ───────────────────────────────────────────────────────────

@app.route("/admin/staff")
@admin_required
def admin_staff():
    # Pass all fields except password hash — phone/email are safe to expose to admin
    safe = [{"username": u,
             "name":  a.get("name", u.title()),
             "role":  a.get("role", "staff"),
             "phone": a.get("phone", ""),
             "email": a.get("email", "")}
            for u, a in _staff_accounts.items()]
    return render_template("staff.html", accounts=safe, dealership=DEALERSHIP_NAME)

@app.route("/admin/staff/add", methods=["POST"])
@admin_required
def admin_staff_add():
    d = request.get_json()
    username = d.get("username","").strip().lower()
    if not username or username in _staff_accounts:
        return jsonify({"ok":False,"message":"Invalid or duplicate username"})
    _staff_accounts[username] = {
        "name":     d.get("name","").strip() or username.title(),
        "role":     d.get("role","staff"),
        "password": d.get("password","changeme"),
        "phone":    d.get("phone","").strip(),
        "email":    d.get("email","").strip(),
    }
    log_activity("staff_add","",f"Account '{username}' created","admin")
    return jsonify({"ok":True,"message":f"Account '{username}' created"})

@app.route("/admin/staff/delete", methods=["POST"])
@admin_required
def admin_staff_delete():
    username = request.get_json().get("username","")
    if username in ("admin","staff"):
        return jsonify({"ok":False,"message":"Cannot delete built-in accounts"})
    if username in _staff_accounts:
        del _staff_accounts[username]
        return jsonify({"ok":True,"message":f"Account '{username}' deleted"})
    return jsonify({"ok":False,"message":"Not found"})

@app.route("/admin/staff/reset-password", methods=["POST"])
@admin_required
def admin_staff_reset():
    d = request.get_json()
    username = d.get("username","")
    new_pw   = d.get("password","").strip()
    if username not in _staff_accounts or not new_pw:
        return jsonify({"ok":False,"message":"Invalid request"})
    _staff_accounts[username]["password"] = new_pw
    log_activity("pw_reset","",f"Password reset for '{username}'","admin")
    return jsonify({"ok":True,"message":"Password updated"})

# ── Existing admin routes ──────────────────────────────────────────────────────

@app.route("/admin/login", methods=["GET","POST"])
def admin_login():
    error = None
    if request.method == "POST":
        if request.form.get("password","") == ADMIN_PASSWORD:
            session.update({"authenticated":True,"is_admin":True,
                            "username":"admin","display_name":"Admin","role":"admin"})
            return redirect(url_for("admin_panel"))
        error = "Invalid admin password."
    return render_template("admin_login.html", error=error)

@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin",None); return redirect(url_for("dashboard"))

@app.route("/admin")
@admin_required
def admin_panel():
    inc = request.args.get("archived")=="1"
    all_v = get_all_vehicles(include_archived=inc)
    sold_m = sold_this_month(all_v)
    return render_template("admin.html",
        vehicles=all_v, statuses=VEHICLE_STATUSES, colours=VEHICLE_COLOURS,
        sold_month=sold_m, monthly_target=_settings.get("monthly_target",15),
        free_tags=free_tags(get_all_vehicles(include_archived=True)),
        suggested_id=next_id(all_v),
        show_archived=inc, recent_log=_activity_log[:20],
        checkouts=list(_key_checkouts.values()),
        aging_90=[v for v in all_v if (v.get("_days_on_lot") or 0) >= 90],
        dealership=DEALERSHIP_NAME)

@app.route("/admin/add", methods=["POST"])
@admin_required
def admin_add():
    fields = form_to_fields(request.form)
    maybe_log_price(fields.get("Internal_ID",""), fields.get("Price"), "admin")
    ok, msg = create_vehicle_record(fields)
    if ok:
        log_activity("add_vehicle", fields.get("Internal_ID",""), "", "admin")
        try:
            _notify_buyers_new_listing(fields)
        except Exception:
            pass
    return jsonify({"ok":ok,"message":msg})

@app.route("/admin/edit/<internal_id>", methods=["GET", "POST"])
@admin_required
def admin_get_vehicle(internal_id):
    v = get_vehicle(internal_id.upper())
    if v: v["_profit"] = calc_profit(v)
    return jsonify({"ok":bool(v),"vehicle":v or {}})

@admin_required
def admin_edit(internal_id):
    fields = form_to_fields(request.form, include_id=False)
    maybe_log_price(internal_id, fields.get("Price"), "admin")
    ok, msg = patch_vehicle(internal_id.upper(), fields, user="admin")
    if ok: log_activity("edit_vehicle", internal_id.upper(), "", "admin")
    return jsonify({"ok":ok,"message":msg})

@app.route("/admin/sell/<internal_id>", methods=["POST"])
@admin_required
def admin_sell(internal_id):
    vid = internal_id.upper()
    sold_fields = {
        "Status": "Sold",
        "Sold_Date": datetime.now().strftime("%Y-%m-%d"),
    }
    d = request.get_json(silent=True) or {}
    buyer         = d.get("buyer","")         or request.form.get("buyer","")
    buyer_dob     = d.get("buyer_dob","")     or request.form.get("buyer_dob","")
    buyer_licence = d.get("buyer_licence","") or request.form.get("buyer_licence","")
    if buyer:         sold_fields["Buyer_Name"]    = buyer.strip()
    if buyer_dob:     sold_fields["Buyer_DOB"]     = buyer_dob.strip()
    if buyer_licence: sold_fields["Buyer_Licence"] = buyer_licence.strip()
    ok, msg = patch_vehicle(vid, sold_fields, user="admin")
    if ok:
        log_activity("sold", vid, "Tag released", "admin")
        vehicle = get_vehicle(vid)
        if vehicle:
            sms_body = (f"SOLD: {vehicle.get('Year','')} {vehicle.get('Make','')} "
                        f"{vehicle.get('Model','')} — {vehicle.get('Price','')} [{vid}]")
            send_sms(MANAGER_PHONE, sms_body)
    return jsonify({"ok":ok,"message":msg})

@app.route("/admin/reassign-tag", methods=["POST"])
@admin_required
def admin_reassign_tag():
    d = request.get_json()
    fi, ti, tag = d.get("from_id","").upper(), d.get("to_id","").upper(), d.get("tag_id","").strip()
    if not all([fi,ti,tag]): return jsonify({"ok":False,"message":"Missing fields"})
    patch_vehicle(fi, {"Key_Tag_ID":""}, user="admin")
    ok, msg = patch_vehicle(ti, {"Key_Tag_ID":tag}, user="admin")
    if ok: log_activity("tag_reassign", ti, f"Tag {tag} from {fi}", "admin")
    return jsonify({"ok":ok,"message":msg})

@app.route("/admin/archive/<internal_id>", methods=["POST"])
@admin_required
def admin_archive(internal_id):
    ok, msg = patch_vehicle(internal_id.upper(), {"Status":"Archived","Key_Tag_ID":None}, user="admin")
    if ok: log_activity("archive", internal_id.upper(), "", "admin")
    return jsonify({"ok":ok,"message":msg})

@app.route("/admin/delete/<internal_id>", methods=["POST"])
@admin_required
def admin_delete(internal_id):
    ok, msg = delete_vehicle_record(internal_id.upper())
    if ok: log_activity("delete", internal_id.upper(), "Permanently deleted", "admin")
    return jsonify({"ok":ok,"message":msg})

# /analytics handled by analytics_page() below

@app.route("/qr-generator")
@login_required
def qr_generator():
    all_v = get_all_vehicles()
    return render_template("qr_generator.html", vehicles=all_v, dealership=DEALERSHIP_NAME)

@app.route("/window-sticker/<internal_id>")
@login_required
def window_sticker(internal_id):
    vehicle = get_vehicle(internal_id.upper())
    if not vehicle: abort(404)
    log_activity("sticker_print", internal_id.upper(), f"{vehicle.get('Make','')} {vehicle.get('Model','')}")
    return render_template("window_sticker.html", vehicle=vehicle,
                           internal_id=internal_id.upper(), dealership=DEALERSHIP_NAME)

@app.route("/bulk-stickers")
@login_required
def bulk_stickers():
    raw = get_all_vehicles()
    # Strip ALL internal/private fields — keep only plain JSON-safe Airtable fields
    PRIVATE = {"_record_id", "_days_on_lot", "_dot_flag", "_profit", "_health"}
    vehicles = []
    for car in raw:
        clean = {}
        for k, v in car.items():
            if k in PRIVATE or k.startswith("_"):
                continue
            # Ensure value is JSON-serialisable
            if isinstance(v, (str, int, float, bool, type(None))):
                clean[k] = v
            else:
                clean[k] = str(v)
        vehicles.append(clean)
    log_activity("bulk_sticker_view", "", str(len(vehicles)) + " vehicles loaded")
    return render_template("bulk_stickers.html",
                           vehicles=vehicles,
                           statuses=VEHICLE_STATUSES,
                           dealership=DEALERSHIP_NAME)

@app.route("/api/analytics-data")
@login_required
def api_analytics_data():
    all_v = get_all_vehicles(include_archived=True)
    status_counts = Counter(v.get("Status","Unknown") for v in all_v)
    make_counts   = Counter(v.get("Make","Unknown") for v in all_v if v.get("Make"))
    return jsonify({
        "ok": True, "total": len(all_v),
        "status_counts": dict(status_counts),
        "make_counts": dict(make_counts.most_common(8)),
        "recent_activity": _activity_log[:30],
        "has_tag": sum(1 for v in all_v if v.get("Key_Tag_ID")),
        "no_tag":  sum(1 for v in all_v if not v.get("Key_Tag_ID") and v.get("Status") not in ("Sold","Archived")),
    })


# ── Inventory Upload ──────────────────────────────────────────────────────────

@app.route("/admin/upload-inventory", methods=["GET"])
@admin_required
def upload_inventory_page():
    all_v = get_all_vehicles(include_archived=True)
    return render_template("upload_inventory.html",
                           dealership=DEALERSHIP_NAME,
                           vehicle_count=len(all_v))

@app.route("/api/upload-inventory", methods=["POST"])
@admin_required
def api_upload_inventory():
    """Smart inventory upload — CSV or Excel. Detects columns, validates, deduplicates."""
    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "message": "No file uploaded"})

    fname = file.filename.lower()
    rows  = []

    # ── Parse file ────────────────────────────────────────────────────────────
    try:
        if fname.endswith(".csv"):
            import csv as _csv, io as _io
            content = file.read().decode("utf-8-sig")
            reader  = _csv.DictReader(_io.StringIO(content))
            rows    = [dict(r) for r in reader]
        elif fname.endswith((".xlsx", ".xls")):
            try:
                import openpyxl as _xl, io as _io
                wb = _xl.load_workbook(_io.BytesIO(file.read()), read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
            except ImportError:
                return jsonify({"ok": False, "message": "Excel support not installed yet — please use CSV for now (openpyxl installs on next deploy)"})
        else:
            return jsonify({"ok": False, "message": "Please upload a .csv or .xlsx file"})
    except Exception as ex:
        return jsonify({"ok": False, "message": f"Could not read file: {ex}"})

    if not rows:
        return jsonify({"ok": False, "message": "File is empty or unreadable"})
    if len(rows) > 300:
        return jsonify({"ok": False, "message": "Maximum 300 rows per upload — split into smaller files"})

    # ── Column normalisation map ───────────────────────────────────────────────
    COL_MAP = {
        # Stock / ID
        "stock#":"Internal_ID","stock":"Internal_ID","stockno":"Internal_ID",
        "stocknumber":"Internal_ID","id":"Internal_ID","stockid":"Internal_ID",
        # VIN
        "vin":"VIN","vinnumber":"VIN","vinno":"VIN","vehicleidentificationnumber":"VIN",
        # Year
        "year":"Year","modelyear":"Year","yr":"Year",
        # Make / Brand
        "make":"Make","brand":"Make","manufacturer":"Make",
        # Model
        "model":"Model",
        # Variant / Trim
        "variant":"Trim","trim":"Trim","trimline":"Trim","trimlevel":"Trim",
        # Series → Notes
        "series":"_series",
        # Body / Category
        "category":"Body","body":"Body","bodystyle":"Body","bodytype":"Body",
        "bodyclass":"Body","vehicletype":"Body",
        # Mileage
        "mileage":"Mileage","odometer":"Mileage","km":"Mileage","kms":"Mileage",
        "kilometers":"Mileage","kilometres":"Mileage","miles":"Mileage",
        # Price
        "price":"Price","saleprice":"Price","askingprice":"Price","listprice":"Price",
        "retailprice":"Price",
        # Status
        "status":"Status","condition":"_condition",
        # Dates
        "listingcreationdate":"Date_Added","dateadded":"Date_Added","date":"Date_Added",
        # Description / Notes
        "description":"Notes","notes":"Notes","comments":"Notes","features":"Notes",
        "remarks":"Notes",
        # Engine
        "engine":"Engine","enginesize":"Engine","enginedescription":"Engine",
        # Transmission
        "transmission":"Transmission","trans":"Transmission","gearbox":"Transmission",
        # Drive
        "drive":"Drive","drivetype":"Drive","drivetrain":"Drive",
        # Fuel
        "fuel":"Fuel","fueltype":"Fuel",
        # Colour
        "colour":"Color","color":"Color","exteriorcolor":"Color","exteriorcolour":"Color",
        "extcolour":"Color","extcolor":"Color",
        # Key tag
        "keytag":"Key_Tag_ID","keytagid":"Key_Tag_ID","keynumber":"Key_Tag_ID",
        # Finance
        "purchaseprice":"Purchase_Price","buyprice":"Purchase_Price","costprice":"Purchase_Price",
        "recon":"Recon_Cost","reconcost":"Recon_Cost","reconditioning":"Recon_Cost",
        # Title (used to extract Year/Make/Model if missing)
        "title":"_title",
    }

    STATUS_MAP = {
        "active":"Available","available":"Available","instock":"Available",
        "for sale":"Available","forsale":"Available","new":"Available",
        "sold":"Sold","deleted":"Archived","inactive":"Archived",
        "archived":"Archived","pending":"Pending","hold":"Hold",
        "demo":"Demo","inservice":"In Service","service":"In Service",
    }

    def norm_col(k):
        return COL_MAP.get(re.sub(r"[^a-z0-9]", "", k.lower().strip()), None)

    def clean_val(v):
        v = str(v).strip()
        return "" if v.lower() in ("none","nan","null","n/a","na","") else v

    def clean_price(v):
        v = re.sub(r"[^\d.]", "", v)
        try:
            p = float(v)
            return f"${p:,.0f}" if p > 0 else ""
        except:
            return ""

    def clean_mileage(v):
        v = re.sub(r"[^\d]", "", v)
        return v if v else ""

    def strip_phone(text):
        # Remove phone numbers — patterns like "0411 764 007" or "o 4 1 1 7 6 4oo7"
        text = re.sub(r'\b0\s*\d[\d\s]{7,12}\b', '', text)
        text = re.sub(r'[oO]\s*[4-9][\s\doO]{8,}', '', text)
        return text.strip()

    def strip_placeholders(text):
        patterns = [
            r'\[Use platform contact options.*?\]',
            r'choose a time that works best for you on \[.*?\]',
            r'Book your inspection time here.*',
        ]
        for p in patterns:
            text = re.sub(p, '', text, flags=re.IGNORECASE)
        return re.sub(r'\n{3,}', '\n\n', text).strip()

    # ── Get existing VINs and IDs for dedup ───────────────────────────────────
    all_v      = get_all_vehicles(include_archived=True)
    exist_vins = {v.get("VIN","").upper().strip() for v in all_v if v.get("VIN")}
    exist_ids  = {v.get("Internal_ID","").upper().strip() for v in all_v if v.get("Internal_ID")}

    results = []
    warnings_total = 0

    for row_idx, raw_row in enumerate(rows, 1):
        # Skip blank rows
        vals = [clean_val(v) for v in raw_row.values()]
        if not any(vals):
            continue

        # Map columns
        mapped = {}
        for raw_key, raw_val in raw_row.items():
            norm = norm_col(raw_key)
            val  = clean_val(raw_val)
            if norm and val:
                mapped[norm] = val

        row_warnings = []

        # ── Extract fields ────────────────────────────────────────────────────
        vin       = mapped.get("VIN","").upper().strip()
        stock_id  = mapped.get("Internal_ID","").upper().strip()
        make      = mapped.get("Make","").title().strip()
        model     = mapped.get("Model","").title().strip()
        year      = mapped.get("Year","").strip()
        price_raw = mapped.get("Price","")
        mileage   = clean_mileage(mapped.get("Mileage",""))
        status_raw= mapped.get("Status","active").lower().strip()
        notes     = mapped.get("Notes","")
        series    = mapped.get("_series","")
        trim      = mapped.get("Trim","").strip()
        body      = mapped.get("Body","").strip()
        date_added= mapped.get("Date_Added","")
        title     = mapped.get("_title","")

        # If no make/model but title like "2018 RENAULT MASTER", parse it
        if title and (not make or not model):
            parts = title.strip().split()
            if parts and parts[0].isdigit() and not year:
                year = parts[0]
            if not make and len(parts) > 1:
                make = parts[1].title()
            if not model and len(parts) > 2:
                model = " ".join(parts[2:]).title()

        # ── VALIDATION ────────────────────────────────────────────────────────

        # VIN
        if not vin:
            row_warnings.append({"level":"error","msg":"Missing VIN"})
        elif len(vin) != 17:
            row_warnings.append({"level":"error","msg":f"VIN wrong length ({len(vin)} chars, needs 17): {vin}"})
        elif not re.match(r'^[A-HJ-NPR-Z0-9]{17}$', vin):
            row_warnings.append({"level":"warn","msg":f"VIN may have invalid characters: {vin}"})

        # Duplicate VIN
        if vin and vin in exist_vins:
            results.append({
                "row": row_idx, "stock": stock_id or vin[:10], "name": f"{year} {make} {model}".strip(),
                "ok": False, "skipped": True,
                "message": f"Already in inventory (VIN {vin}) — skipped",
                "warnings": []
            })
            continue

        # Duplicate Stock ID
        if stock_id and stock_id in exist_ids:
            results.append({
                "row": row_idx, "stock": stock_id, "name": f"{year} {make} {model}".strip(),
                "ok": False, "skipped": True,
                "message": f"Stock # {stock_id} already exists — skipped",
                "warnings": []
            })
            continue

        # Year
        if year:
            try:
                y = int(year)
                if y < 1980 or y > 2027:
                    row_warnings.append({"level":"warn","msg":f"Year looks unusual: {year}"})
            except:
                row_warnings.append({"level":"warn","msg":f"Year is not a number: {year}"})

        # Price
        price = clean_price(price_raw)
        if price_raw and not price:
            row_warnings.append({"level":"warn","msg":f"Could not parse price: {price_raw}"})
        elif price:
            try:
                p = float(re.sub(r"[^\d.]", "", price))
                if p < 500:
                    row_warnings.append({"level":"warn","msg":f"Price very low: {price}"})
                elif p > 500000:
                    row_warnings.append({"level":"warn","msg":f"Price very high: {price}"})
            except:
                pass

        # Mileage
        if mileage:
            try:
                m = int(mileage)
                if m > 500000:
                    row_warnings.append({"level":"warn","msg":f"Mileage very high: {m:,}km"})
            except:
                row_warnings.append({"level":"warn","msg":f"Mileage not a number: {mileage}"})

        # Missing critical
        if not make:
            row_warnings.append({"level":"error","msg":"Missing Make"})
        if not model:
            row_warnings.append({"level":"warn","msg":"Missing Model"})

        # Clean description
        if notes:
            original = notes
            notes = strip_phone(notes)
            notes = strip_placeholders(notes)
            if series:
                notes = f"{series}\n\n{notes}".strip()
            if notes != original:
                row_warnings.append({"level":"info","msg":"Description cleaned (phone/placeholder removed)"})

        # Status mapping
        status = STATUS_MAP.get(re.sub(r"[^a-z]","",status_raw), "Available")

        # ── Build fields ──────────────────────────────────────────────────────
        if not stock_id:
            all_v = get_all_vehicles(include_archived=True)
            exist_ids_live = {v.get("Internal_ID","").upper() for v in all_v}
            # find next available ASR or CAR id
            prefix = "ASR" if any(v.get("Internal_ID","").startswith("ASR") for v in all_v) else "CAR"
            stock_id = next_id(all_v, prefix)

        fields = {"Internal_ID": stock_id, "Status": status,
                  "Date_Added": date_added or datetime.now(timezone.utc).strftime("%Y-%m-%d")}

        for key, val in [
            ("VIN",vin),("Make",make),("Model",model),
            ("Year", int(year) if year and str(year).isdigit() else year),
            ("Trim",trim),("Body",body),("Price",price),
            ("Mileage", f"{int(mileage):,}km" if mileage and mileage.isdigit() else mileage),
            ("Notes",notes),
            ("Engine",   mapped.get("Engine","")),
            ("Transmission", mapped.get("Transmission","")),
            ("Drive",    mapped.get("Drive","")),
            ("Fuel",     mapped.get("Fuel","")),
            ("Color",    mapped.get("Color","")),
            ("Key_Tag_ID", mapped.get("Key_Tag_ID","")),
            ("Purchase_Price", mapped.get("Purchase_Price","")),
            ("Recon_Cost",     mapped.get("Recon_Cost","")),
        ]:
            if val and val not in ("","0",0):
                fields[key] = val

        # Check for errors that block upload
        has_error = any(w["level"]=="error" for w in row_warnings)
        if has_error:
            results.append({
                "row": row_idx, "stock": stock_id, "name": f"{year} {make} {model}".strip(),
                "ok": False, "skipped": False,
                "message": "Not uploaded — fix errors first",
                "warnings": row_warnings
            })
            warnings_total += 1
            continue

        # ── Upload to Airtable ────────────────────────────────────────────────
        ok, msg = create_vehicle_record(fields)
        if ok:
            exist_vins.add(vin)
            exist_ids.add(stock_id)
            all_v.append(fields)

        results.append({
            "row": row_idx, "stock": stock_id,
            "name": f"{year} {make} {model}".strip(),
            "ok": ok, "skipped": False,
            "message": "Added to inventory" if ok else f"Airtable error: {msg}",
            "warnings": row_warnings
        })
        if row_warnings:
            warnings_total += 1

    created = sum(1 for r in results if r["ok"])
    skipped = sum(1 for r in results if r.get("skipped"))
    failed  = sum(1 for r in results if not r["ok"] and not r.get("skipped"))

    return jsonify({
        "ok": True,
        "results": results,
        "created": created,
        "skipped": skipped,
        "failed":  failed,
        "warnings": warnings_total,
    })


# ══════════════════════════════════════════════════════════════════════════════
# BATCH 1 — Finance, Pipeline, Health Score, AI Listing, Salesperson Mode
# ══════════════════════════════════════════════════════════════════════════════

# ── Finance Calculator ────────────────────────────────────────────────────────

@app.route("/finance")
@login_required
def finance_page():
    all_v = get_all_vehicles()
    avail = [v for v in all_v if v.get("Status") == "Available"]
    return render_template("finance.html", vehicles=avail, dealership=DEALERSHIP_NAME)

@app.route("/api/finance-calc", methods=["POST"])
@login_required
def api_finance_calc():
    d = request.get_json(silent=True) or {}
    vid      = d.get("vid","").upper().strip()
    deposit  = float(re.sub(r"[^\d.]","", str(d.get("deposit","0"))) or 0)
    term     = int(d.get("term_years", 5))
    vehicle  = get_vehicle(vid) if vid else {"Price": str(d.get("price","0"))}
    if not vehicle:
        return jsonify({"ok": False, "message": "Vehicle not found"})
    result = gen_finance_options(vehicle, deposit=deposit, term_years=term)
    if not result:
        return jsonify({"ok": False, "message": "Could not parse vehicle price"})
    return jsonify({"ok": True, **result})

# ── Deal Pipeline ─────────────────────────────────────────────────────────────

PIPELINE_STAGES = ["Lead","Contacted","Test Drive","Negotiation","Deposit","Sold","Lost"]

@app.route("/pipeline")
@login_required
def pipeline_page():
    all_v = get_all_vehicles()
    vid_map = {v.get("Internal_ID","").upper(): v for v in all_v}
    # Group cards by stage
    board = {s: [] for s in PIPELINE_STAGES}
    for card in _pipeline:
        stage = card.get("stage","Lead")
        if stage in board:
            v = vid_map.get(card.get("vid","").upper(), {})
            board[stage].append({**card, "_vehicle": v})
    return render_template("pipeline.html", board=board, stages=PIPELINE_STAGES,
                           vehicles=all_v, dealership=DEALERSHIP_NAME)

@app.route("/api/pipeline/add", methods=["POST"])
@login_required
def api_pipeline_add():
    d = request.get_json(silent=True) or {}
    vid = d.get("vid","").upper().strip()
    if not vid:
        return jsonify({"ok": False, "message": "Vehicle ID required"})
    if any(c["vid"] == vid for c in _pipeline):
        return jsonify({"ok": False, "message": "Vehicle already in pipeline"})
    card = {
        "id":       vid + "_" + str(uuid.uuid4())[:6].upper(),
        "vid":      vid,
        "stage":    d.get("stage","Lead"),
        "customer": d.get("customer",""),
        "phone":    d.get("phone",""),
        "notes":    d.get("notes",""),
        "created":  datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"),
        "updated":  datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"),
        "staff":    current_user(),
    }
    _pipeline.append(card)
    log_activity("pipeline_add", vid, "Added to pipeline at " + card["stage"], current_user())
    return jsonify({"ok": True, "card": card})

@app.route("/api/pipeline/move", methods=["POST"])
@login_required
def api_pipeline_move():
    d = request.get_json(silent=True) or {}
    card_id = d.get("id","")
    new_stage = d.get("stage","")
    if new_stage not in PIPELINE_STAGES:
        return jsonify({"ok": False, "message": "Invalid stage"})
    for card in _pipeline:
        if card["id"] == card_id:
            old_stage = card["stage"]
            card["stage"] = new_stage
            card["updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
            log_activity("pipeline_move", card["vid"],
                         old_stage + " → " + new_stage, current_user())
            if new_stage == "Sold":
                patch_vehicle(card["vid"], {"Status":"Sold",
                    "Sold_Date": datetime.now().strftime("%Y-%m-%d")}, user=current_user())
            return jsonify({"ok": True})
    return jsonify({"ok": False, "message": "Card not found"})

@app.route("/api/pipeline/update", methods=["POST"])
@login_required
def api_pipeline_update():
    d = request.get_json(silent=True) or {}
    card_id = d.get("id","")
    for card in _pipeline:
        if card["id"] == card_id:
            for k in ("customer","phone","notes","stage"):
                if k in d: card[k] = d[k]
            card["updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
            return jsonify({"ok": True})
    return jsonify({"ok": False, "message": "Card not found"})

@app.route("/api/pipeline/remove", methods=["POST"])
@admin_required
def api_pipeline_remove():
    d = request.get_json(silent=True) or {}
    card_id = d.get("id","")
    global _pipeline
    before = len(_pipeline)
    _pipeline = [c for c in _pipeline if c["id"] != card_id]
    return jsonify({"ok": len(_pipeline) < before})

@app.route("/api/pipeline/all")
@login_required
def api_pipeline_all():
    return jsonify({"ok": True, "pipeline": _pipeline, "stages": PIPELINE_STAGES})

# ── Inventory Health ──────────────────────────────────────────────────────────

@app.route("/inventory-health")
@login_required
def inventory_health_page():
    all_v = get_all_vehicles()
    scored = []
    for v in all_v:
        if v.get("Status") in ("Sold","Archived"): continue
        hs = health_score(v)
        scored.append({**v, "_health": hs})
    scored.sort(key=lambda x: x["_health"]["score"])
    avg = round(sum(x["_health"]["score"] for x in scored) / len(scored)) if scored else 0
    grade_counts = {"A":0,"B":0,"C":0,"D":0}
    for x in scored:
        grade_counts[x["_health"]["grade"]] += 1
    return render_template("inventory_health.html", vehicles=scored, avg_score=avg,
                           grade_counts=grade_counts, dealership=DEALERSHIP_NAME)

@app.route("/api/health-score/<vid>")
@login_required
def api_health_score(vid):
    v = get_vehicle(vid.upper())
    if not v: return jsonify({"ok": False, "message": "Not found"})
    return jsonify({"ok": True, **health_score(v)})

# ── AI Listing Generator ──────────────────────────────────────────────────────

@app.route("/ai-listings")
@login_required
def ai_listings_page():
    all_v = get_all_vehicles()
    avail = [v for v in all_v if v.get("Status") not in ("Sold","Archived")]
    return render_template("ai_listings.html", vehicles=avail, dealership=DEALERSHIP_NAME)

@app.route("/api/gen-listing", methods=["POST"])
@login_required
def api_gen_listing():
    d = request.get_json(silent=True) or {}
    vid      = d.get("vid","").upper().strip()
    platform = d.get("platform","carsales")
    tone     = d.get("tone","standard")
    vehicle  = get_vehicle(vid)
    if not vehicle:
        return jsonify({"ok": False, "message": "Vehicle not found"})
    text = gen_listing_full(vehicle, platform=platform, tone=tone)
    return jsonify({"ok": True, "text": text, "vid": vid,
                    "platform": platform, "tone": tone})

# ── Salesperson Mode ──────────────────────────────────────────────────────────

@app.route("/api/salesperson-brief/<vid>")
@login_required
def api_salesperson_brief(vid):
    vehicle = get_vehicle(vid.upper())
    if not vehicle:
        return jsonify({"ok": False, "message": "Vehicle not found"})
    brief = gen_salesperson_brief(vehicle)
    finance = gen_finance_options(vehicle, deposit=0, term_years=5)
    return jsonify({"ok": True, "brief": brief, "finance": finance, "vehicle": {
        "vid": vid, "name": str(vehicle.get("Year","")) + " " +
               str(vehicle.get("Make","")) + " " + str(vehicle.get("Model","")),
        "price": vehicle.get("Price",""), "mileage": vehicle.get("Mileage",""),
    }})

# ── Trade-In Manager ──────────────────────────────────────────────────────────

@app.route("/trade-ins")
@login_required
def trade_ins_page():
    all_v = get_all_vehicles()
    return render_template("trade_ins.html", trade_ins=_trade_ins,
                           vehicles=all_v, dealership=DEALERSHIP_NAME)

@app.route("/api/trade-in/add", methods=["POST"])
@login_required
def api_trade_in_add():
    d = request.get_json(silent=True) or {}
    required = ["customer","vin_trade","make","model","year"]
    for r in required:
        if not d.get(r):
            return jsonify({"ok": False, "message": "Missing: " + r})
    tid = "TI-" + str(uuid.uuid4())[:8].upper()
    trade = {
        "id":          tid,
        "customer":    d.get("customer",""),
        "phone":       d.get("phone",""),
        "vin_trade":   d.get("vin_trade","").upper(),
        "make":        d.get("make",""),
        "model":       d.get("model",""),
        "year":        d.get("year",""),
        "color":       d.get("color",""),
        "mileage":     d.get("mileage",""),
        "condition":   d.get("condition","Good"),
        "offer_price": d.get("offer_price",""),
        "notes":       d.get("notes",""),
        "status":      "Pending",
        "created":     datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"),
        "staff":       current_user(),
        "vehicle_interest": d.get("vehicle_interest",""),
    }
    _trade_ins.append(trade)
    log_activity("trade_in", tid, "Trade-in added: " + trade["year"] + " " + trade["make"] + " " + trade["model"], current_user())
    return jsonify({"ok": True, "trade": trade})

@app.route("/api/trade-in/convert/<tid>", methods=["POST"])
@admin_required
def api_trade_in_convert(tid):
    """Convert a trade-in into an inventory vehicle."""
    trade = next((t for t in _trade_ins if t["id"] == tid), None)
    if not trade:
        return jsonify({"ok": False, "message": "Trade-in not found"})
    all_v = get_all_vehicles(include_archived=True)
    new_id = next_id(all_v, "TRD")
    fields = {
        "Internal_ID": new_id,
        "VIN":    trade["vin_trade"],
        "Make":   trade["make"],
        "Model":  trade["model"],
        "Year":   int(trade["year"]) if str(trade["year"]).isdigit() else trade["year"],
        "Color":  trade.get("color",""),
        "Mileage":trade.get("mileage",""),
        "Status": "Available",
        "Notes":  "Trade-in from " + trade["customer"] + ". Condition: " + trade.get("condition","") + ". " + trade.get("notes",""),
        "Purchase_Price": _to_num(trade.get("offer_price","")),
        "Date_Added": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
    }
    fields = {k:v for k,v in fields.items() if v}
    ok, msg = create_vehicle_record(fields)
    if ok:
        trade["status"] = "Converted"
        trade["inventory_id"] = new_id
    return jsonify({"ok": ok, "message": msg, "inventory_id": new_id if ok else ""})

@app.route("/api/trade-in/update", methods=["POST"])
@login_required
def api_trade_in_update():
    d = request.get_json(silent=True) or {}
    tid = d.get("id","")
    trade = next((t for t in _trade_ins if t["id"] == tid), None)
    if not trade:
        return jsonify({"ok": False, "message": "Not found"})
    for k in ("offer_price","condition","notes","status"):
        if k in d: trade[k] = d[k]
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# BATCH 2 — Advanced Analytics, Global Search, Document Generator
# ══════════════════════════════════════════════════════════════════════════════

# ── Advanced Analytics KPIs ───────────────────────────────────────────────────

@app.route("/analytics")
@login_required
def analytics_page():
    all_v = get_all_vehicles(include_archived=True)
    now   = datetime.now()

    # ── Sales & revenue ───────────────────────────────────────────────────────
    sold_all = [v for v in all_v if v.get("Status") == "Sold"]
    sold_this_m = sold_this_month(all_v)

    # Monthly sales last 6 months
    monthly_sales = {}
    for v in sold_all:
        sd = v.get("Sold_Date","")
        if sd:
            try:
                d = datetime.fromisoformat(sd[:10])
                key = d.strftime("%b %Y")
                monthly_sales[key] = monthly_sales.get(key, 0) + 1
            except Exception:
                pass
    # last 6 months ordered
    from calendar import month_abbr
    months_ordered = []
    for i in range(5, -1, -1):
        # Correctly go back i months from now
        month_num = now.month - i
        year_offset = 0
        while month_num <= 0:
            month_num += 12
            year_offset -= 1
        y = now.year + year_offset
        key = month_abbr[month_num] + " " + str(y)
        months_ordered.append({"label": key, "count": monthly_sales.get(key, 0)})

    # ── Profit & margin ───────────────────────────────────────────────────────
    profits_by_make = defaultdict(list)
    for v in sold_all:
        p = calc_profit(v)
        if p["has_data"] and p["profit"] is not None:
            profits_by_make[v.get("Make","Unknown")].append(p["profit"])

    make_profit = sorted(
        [{"make": k, "avg_profit": round(sum(v)/len(v)), "count": len(v), "total": round(sum(v))}
         for k, v in profits_by_make.items()],
        key=lambda x: x["avg_profit"], reverse=True
    )[:8]

    all_profits = [p for v in sold_all for p_ in [calc_profit(v)] if p_["has_data"] and p_["profit"] for p in [p_["profit"]]]
    total_profit  = round(sum(all_profits)) if all_profits else 0
    avg_profit    = round(sum(all_profits)/len(all_profits)) if all_profits else 0
    total_revenue = round(sum(
        float(re.sub(r"[^\d.]","",str(v.get("Price","") or "")) or 0) for v in sold_all
    ))

    # ── Days to sell ──────────────────────────────────────────────────────────
    days_to_sell = []
    for v in sold_all:
        try:
            added = datetime.fromisoformat(v.get("Date_Added","")[:10])
            sold  = datetime.fromisoformat(v.get("Sold_Date","")[:10])
            days_to_sell.append((sold - added).days)
        except Exception:
            pass
    avg_days_to_sell = round(sum(days_to_sell)/len(days_to_sell)) if days_to_sell else 0

    # ── Per-make days to sell ─────────────────────────────────────────────────
    make_days = defaultdict(list)
    for v in sold_all:
        try:
            added = datetime.fromisoformat(v.get("Date_Added","")[:10])
            sold  = datetime.fromisoformat(v.get("Sold_Date","")[:10])
            make_days[v.get("Make","Unknown")].append((sold-added).days)
        except Exception:
            pass
    make_velocity = sorted(
        [{"make": k, "avg_days": round(sum(v)/len(v)), "count": len(v)}
         for k, v in make_days.items() if len(v) >= 1],
        key=lambda x: x["avg_days"]
    )[:8]

    # ── Staff performance ─────────────────────────────────────────────────────
    staff_sales = Counter(v.get("Buyer_Name","") or "Unknown" for v in sold_all)
    staff_enquiries = Counter(e.get("staff","Unknown") for e in _enquiries)
    staff_test_drives = Counter(td.get("staff","Unknown") for td in _test_drives)

    # Conversion rate: test drives → sales (by vehicle)
    td_vids  = set(td["internal_id"] for td in _test_drives)
    sold_vids = set(v.get("Internal_ID","") for v in sold_all)
    conversion_rate = round(len(td_vids & sold_vids) / len(td_vids) * 100) if td_vids else 0

    # ── Inventory health summary ──────────────────────────────────────────────
    active_v = [v for v in all_v if v.get("Status") not in ("Sold","Archived")]
    health_scores = [health_score(v)["score"] for v in active_v]
    avg_health = round(sum(health_scores)/len(health_scores)) if health_scores else 0
    slow_movers = [v for v in active_v if (v.get("_days_on_lot") or 0) >= 60]

    # ── Make breakdown ────────────────────────────────────────────────────────
    make_breakdown = Counter(v.get("Make","Unknown") for v in active_v if v.get("Make")).most_common(10)

    # ── Status summary ────────────────────────────────────────────────────────
    status_counts = Counter(v.get("Status","Unknown") for v in all_v)

    return render_template("analytics.html",
        dealership=DEALERSHIP_NAME,
        all_vehicles=all_v, active_v=active_v,
        sold_count=len(sold_all), sold_this_month=sold_this_m,
        monthly_sales=months_ordered,
        total_revenue=total_revenue, total_profit=total_profit, avg_profit=avg_profit,
        make_profit=make_profit, make_velocity=make_velocity,
        avg_days_to_sell=avg_days_to_sell, conversion_rate=conversion_rate,
        staff_sales=dict(staff_sales.most_common(10)),
        staff_enquiries=dict(staff_enquiries.most_common(10)),
        staff_test_drives=dict(staff_test_drives.most_common(10)),
        avg_health=avg_health, slow_movers=len(slow_movers),
        make_breakdown=dict(make_breakdown),
        status_counts=dict(status_counts),
        enquiry_count=len(_enquiries), test_drive_count=len(_test_drives),
        pipeline_count=len(_pipeline),
    )

# ── Global Search ─────────────────────────────────────────────────────────────

@app.route("/search")
@login_required
def search_page():
    return render_template("search.html", dealership=DEALERSHIP_NAME)

@app.route("/api/search")
@login_required
def api_search():
    q = request.args.get("q","").strip().lower()
    if len(q) < 2:
        return jsonify({"ok": True, "results": [], "total": 0})

    results = []

    # ── Search vehicles ───────────────────────────────────────────────────────
    all_v = get_all_vehicles(include_archived=True)
    for v in all_v:
        score = 0
        fields_to_search = [
            v.get("Internal_ID",""), v.get("VIN",""), v.get("Make",""),
            v.get("Model",""), v.get("Trim",""), v.get("Color",""),
            v.get("Notes",""), v.get("Engine",""), str(v.get("Year","")),
            v.get("Buyer_Name",""),
        ]
        haystack = " ".join(str(f) for f in fields_to_search).lower()
        if q in haystack:
            # Boost exact matches on key fields
            if q == v.get("Internal_ID","").lower(): score += 100
            elif q == v.get("VIN","").lower(): score += 90
            elif q in v.get("Internal_ID","").lower(): score += 50
            else: score += 10
            results.append({
                "type": "vehicle",
                "id": v.get("Internal_ID",""),
                "title": str(v.get("Year","")) + " " + str(v.get("Make","")) + " " + str(v.get("Model","")),
                "subtitle": v.get("Internal_ID","") + " · " + (v.get("VIN","")[:10] or "No VIN") + " · " + (v.get("Status","") or ""),
                "meta": (v.get("Price","") or "") + " · " + (v.get("Color","") or "") + " · " + (str(v.get("_days_on_lot","")) + " days" if v.get("_days_on_lot") is not None else ""),
                "url": "/scan/" + v.get("Internal_ID",""),
                "status": v.get("Status",""),
                "score": score,
            })

    # ── Search customers (enquiries) ──────────────────────────────────────────
    seen_customers = set()
    for e in _enquiries:
        ckey = e.get("name","").lower()
        haystack = (e.get("name","") + " " + e.get("phone","") + " " + e.get("email","")).lower()
        if q in haystack and ckey not in seen_customers:
            seen_customers.add(ckey)
            results.append({
                "type": "customer",
                "id": e.get("id",""),
                "title": e.get("name",""),
                "subtitle": (e.get("phone","") or "No phone") + " · " + (e.get("email","") or ""),
                "meta": "Enquired about " + e.get("internal_id","") + " · " + e.get("timestamp",""),
                "url": "/customers",
                "status": e.get("follow_up_status",""),
                "score": 30,
            })

    # ── Search pipeline ───────────────────────────────────────────────────────
    for card in _pipeline:
        haystack = (card.get("customer","") + " " + card.get("phone","") + " " + card.get("vid","") + " " + card.get("notes","")).lower()
        if q in haystack:
            results.append({
                "type": "pipeline",
                "id": card.get("id",""),
                "title": card.get("customer","Unknown") + " — " + card.get("vid",""),
                "subtitle": "Stage: " + card.get("stage","") + " · " + card.get("updated",""),
                "meta": card.get("notes","")[:80],
                "url": "/pipeline",
                "status": card.get("stage",""),
                "score": 20,
            })

    # ── Search trade-ins ──────────────────────────────────────────────────────
    for t in _trade_ins:
        haystack = (t.get("customer","") + " " + t.get("vin_trade","") + " " +
                    t.get("make","") + " " + t.get("model","") + " " + t.get("phone","")).lower()
        if q in haystack:
            results.append({
                "type": "trade-in",
                "id": t.get("id",""),
                "title": t.get("customer","") + " — " + str(t.get("year","")) + " " + t.get("make","") + " " + t.get("model",""),
                "subtitle": "Trade-in · Offer: " + str(t.get("offer_price","—")) + " · " + t.get("status",""),
                "meta": "VIN: " + (t.get("vin_trade","") or "—") + " · " + t.get("condition",""),
                "url": "/trade-ins",
                "status": t.get("status",""),
                "score": 20,
            })

    results.sort(key=lambda x: x["score"], reverse=True)
    return jsonify({"ok": True, "results": results[:40], "total": len(results), "query": q})

# ── Document Generator ────────────────────────────────────────────────────────

@app.route("/documents")
@login_required
def documents_page():
    all_v = get_all_vehicles()
    return render_template("documents.html", vehicles=all_v, dealership=DEALERSHIP_NAME)

@app.route("/api/document/<doc_type>/<vid>")
@login_required
def api_generate_document(doc_type, vid):
    vehicle = get_vehicle(vid.upper())
    if not vehicle:
        return "Vehicle not found", 404
    now_str  = datetime.now().strftime("%d %B %Y")
    now_time = datetime.now().strftime("%d/%m/%Y %H:%M")
    staff    = session.get("display_name", current_user())
    dealership = _settings.get("dealership_name", DEALERSHIP_NAME)
    v = vehicle
    name  = str(v.get("Year","")) + " " + str(v.get("Make","")) + " " + str(v.get("Model",""))
    price = v.get("Price","—")
    vin   = v.get("VIN","—")
    vid   = v.get("Internal_ID","")

    if doc_type == "sales-agreement":
        buyer_name = v.get("Buyer_Name","_________________________")
        buyer_dob  = v.get("Buyer_DOB","___________")
        buyer_lic  = v.get("Buyer_Licence","___________")
        return render_template("doc_sales_agreement.html",
            vehicle=v, name=name, price=price, vin=vin, vid=vid,
            buyer_name=buyer_name, buyer_dob=buyer_dob, buyer_lic=buyer_lic,
            date=now_str, staff=staff, dealership=dealership)

    elif doc_type == "test-drive-waiver":
        return render_template("doc_test_drive_waiver.html",
            vehicle=v, name=name, vin=vin, vid=vid,
            date=now_str, time=now_time, staff=staff, dealership=dealership)

    elif doc_type == "deposit-receipt":
        return render_template("doc_deposit_receipt.html",
            vehicle=v, name=name, price=price, vin=vin, vid=vid,
            date=now_str, staff=staff, dealership=dealership)

    elif doc_type == "vehicle-inspection":
        return render_template("doc_vehicle_inspection.html",
            vehicle=v, name=name, vin=vin, vid=vid,
            date=now_str, staff=staff, dealership=dealership)

    return "Document type not found", 404


# ══════════════════════════════════════════════════════════════════════════════
# BATCH 3 — Role Permissions, Inline Editing, Global Search shortcuts
# ══════════════════════════════════════════════════════════════════════════════

ROLE_PERMISSIONS = {
    "admin":   {"can_edit","can_delete","can_sell","can_view","can_add","can_assign_keys","can_manage_staff","can_export"},
    "manager": {"can_edit","can_sell","can_view","can_add","can_assign_keys","can_export"},
    "sales":   {"can_edit","can_view","can_add","can_assign_keys"},
    "service": {"can_view","can_assign_keys"},
    "viewer":  {"can_view"},
    "staff":   {"can_edit","can_view","can_add","can_assign_keys"},  # default backwards-compat
}

def has_permission(perm):
    role = session.get("role", "staff")
    return perm in ROLE_PERMISSIONS.get(role, set())

def permission_required(perm):
    def decorator(f):
        @wraps(f)
        def d(*a, **kw):
            if not is_auth():
                return redirect(url_for("login"))
            if not has_permission(perm):
                if request.method == "POST" or request.headers.get("X-Requested-With"):
                    return jsonify({"ok":False,"message":"Permission denied — your role cannot perform this action"}), 403
                return render_template("403.html"), 403
            return f(*a, **kw)
        return d
    return decorator

# ── Inline Edit API ────────────────────────────────────────────────────────────

INLINE_EDITABLE = {
    "Price":        {"type":"text","label":"Price"},
    "Status":       {"type":"select","label":"Status","options":["Available","Hold","Demo","In Service","Pending","Sold","Archived"]},
    "Mileage":      {"type":"text","label":"Mileage"},
    "Color":        {"type":"text","label":"Colour"},
    "Notes":        {"type":"textarea","label":"Notes"},
    "Key_Tag_ID":   {"type":"text","label":"Key Tag ID"},
    "Recon_Cost":   {"type":"text","label":"Recon Cost"},
    "Purchase_Price":{"type":"text","label":"Purchase Price"},
}

@app.route("/api/inline-edit", methods=["POST"])
@login_required
def api_inline_edit():
    d = request.get_json(silent=True) or {}
    vid   = d.get("vid","").upper().strip()
    field = d.get("field","").strip()
    value = d.get("value","")

    if not vid or not field:
        return jsonify({"ok":False,"message":"vid and field required"})
    if field not in INLINE_EDITABLE:
        return jsonify({"ok":False,"message":"Field not editable inline: " + field})

    # Permission check
    role = session.get("role","staff")
    # Status changes and sell require elevated perms
    if field in ("Status",) and not has_permission("can_edit"):
        return jsonify({"ok":False,"message":"Permission denied"})

    maybe_log_price(vid, value if field == "Price" else None, current_user())
    # Cast numeric fields to float before sending to Airtable
    FLOAT_FIELDS = {"Price", "Purchase_Price", "Recon_Cost"}
    if field in FLOAT_FIELDS:
        value = _to_num(value)
        if value is None:
            return jsonify({"ok": False, "message": "Please enter a valid number"})
    ok, msg = patch_vehicle(vid, {field: value}, user=current_user())
    if ok:
        log_activity("inline_edit", vid, field + " → " + str(value)[:40], current_user())
    return jsonify({"ok":ok,"message":msg,"field":field,"value":value,"vid":vid})

@app.route("/api/inline-editable-fields")
@login_required
def api_inline_editable_fields():
    return jsonify({"ok":True,"fields":INLINE_EDITABLE,"statuses":VEHICLE_STATUSES})

# ── Session / Role Info API ───────────────────────────────────────────────────

@app.route("/api/session-info")
@login_required
def api_session_info():
    role = session.get("role","staff")
    return jsonify({
        "ok": True,
        "username":     current_user(),
        "display_name": session.get("display_name","Staff"),
        "role":         role,
        "is_admin":     session.get("is_admin", False),
        "permissions":  list(ROLE_PERMISSIONS.get(role, set())),
    })

# ── Command Palette suggestions API ──────────────────────────────────────────

@app.route("/api/cmd-palette")
@login_required
def api_cmd_palette():
    q = request.args.get("q","").strip().lower()
    nav_items = [
        {"label":"Dashboard",        "icon":"🚗","url":"/dashboard"},
        {"label":"Analytics",        "icon":"📈","url":"/analytics"},
        {"label":"Pipeline",         "icon":"🎯","url":"/pipeline"},
        {"label":"Finance Calculator","icon":"💰","url":"/finance"},
        {"label":"Inventory Health", "icon":"🏥","url":"/inventory-health"},
        {"label":"AI Listings",      "icon":"🤖","url":"/ai-listings"},
        {"label":"Trade-Ins",        "icon":"🔄","url":"/trade-ins"},
        {"label":"Search",           "icon":"🔍","url":"/search"},
        {"label":"Documents",        "icon":"📄","url":"/documents"},
        {"label":"Customers",        "icon":"👥","url":"/customers"},
        {"label":"Key Board",        "icon":"🔑","url":"/keyboard"},
        {"label":"Scanner",          "icon":"📷","url":"/scanner"},
        {"label":"Compare",          "icon":"⚖️","url":"/compare"},
        {"label":"QR Codes",         "icon":"📲","url":"/qr-generator"},
        {"label":"Reports",          "icon":"📊","url":"/reports"},
        {"label":"Analytics",        "icon":"📈","url":"/analytics"},
    ]
    if session.get("is_admin"):
        nav_items += [
            {"label":"Admin Panel",  "icon":"⚙️","url":"/admin"},
            {"label":"Audit Trail",  "icon":"📋","url":"/admin/audit"},
            {"label":"Staff Manager","icon":"👤","url":"/admin/staff"},
            {"label":"Settings",     "icon":"🔧","url":"/settings"},
            {"label":"Bulk Upload",  "icon":"⬆️","url":"/admin/upload-inventory"},
        ]
    if not q:
        return jsonify({"ok":True,"results":nav_items[:10]})
    # Filter by query
    filtered = [x for x in nav_items if q in x["label"].lower()]
    # Also search vehicles
    if len(q) >= 2:
        veh_results = []
        for v in get_all_vehicles():
            haystack = (str(v.get("Internal_ID","")) + " " + str(v.get("Make","")) + " " +
                        str(v.get("Model","")) + " " + str(v.get("VIN",""))).lower()
            if q in haystack:
                veh_results.append({
                    "label": str(v.get("Year","")) + " " + str(v.get("Make","")) + " " + str(v.get("Model","")) + " (" + str(v.get("Internal_ID","")) + ")",
                    "icon": "🚗",
                    "url": "/scan/" + str(v.get("Internal_ID","")),
                    "sublabel": v.get("Status","") + " · " + str(v.get("Price",""))
                })
                if len(veh_results) >= 5: break
        filtered = filtered + veh_results
    return jsonify({"ok":True,"results":filtered[:12]})


@app.route("/roles")
@admin_required
def roles_page():
    return render_template("roles.html", dealership=DEALERSHIP_NAME)


# ── Bulk Import (Excel/CSV via bulk_import.html template) ─────────────────────

@app.route("/admin/bulk-import")
@admin_required
def bulk_import_page():
    return render_template("bulk_import.html", dealership=DEALERSHIP_NAME)

@app.route("/api/bulk-import-excel", methods=["POST"])
@admin_required
def api_bulk_import_excel():
    """Process Excel/CSV upload from bulk_import.html template."""
    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "message": "No file uploaded"})

    filename = file.filename.lower()
    rows = []
    try:
        if filename.endswith(".csv"):
            import csv, io
            content = file.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif filename.endswith((".xlsx", ".xls")):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v or "").strip() for v in row])))
            except ImportError:
                return jsonify({"ok": False, "message": "openpyxl not installed — use CSV format"})
        else:
            return jsonify({"ok": False, "message": "Only CSV or XLSX files accepted"})
    except Exception as e:
        return jsonify({"ok": False, "message": "File read error: " + str(e)})

    if not rows:
        return jsonify({"ok": False, "message": "File is empty or could not be parsed"})

    # ── Column mapping (same as upload_inventory) ─────────────────────────────
    COL_MAP = {
        "stock #": "Internal_ID", "stock#": "Internal_ID", "stock no": "Internal_ID",
        "internal_id": "Internal_ID", "id": "Internal_ID",
        "vin": "VIN", "vin number": "VIN",
        "year": "Year", "make": "Make", "model": "Model",
        "variant": "Trim", "trim": "Trim", "series": "Trim",
        "category": "Body", "body": "Body", "body type": "Body",
        "mileage": "Mileage", "odometer": "Mileage", "km": "Mileage",
        "price": "Price", "asking price": "Price", "sale price": "Price",
        "status": "Status", "listing status": "Status",
        "colour": "Color", "color": "Color",
        "engine": "Engine", "transmission": "Transmission",
        "drive": "Drive", "fuel": "Fuel", "fuel type": "Fuel",
        "notes": "Notes", "description": "Notes", "comments": "Notes",
        "date added": "Date_Added", "listing creation date": "Date_Added",
        "purchase price": "Purchase_Price", "cost price": "Purchase_Price",
        "recon": "Recon_Cost", "recon cost": "Recon_Cost",
    }
    STATUS_MAP = {
        "active": "Available", "live": "Available", "for sale": "Available",
        "sold": "Sold", "hold": "Hold", "demo": "Demo",
        "in service": "In Service", "pending": "Pending", "archived": "Archived",
    }

    # Existing VINs and IDs for dedup
    existing = get_all_vehicles(include_archived=True)
    existing_vins = {v.get("VIN","").upper() for v in existing if v.get("VIN")}
    existing_ids  = {v.get("Internal_ID","").upper() for v in existing if v.get("Internal_ID")}

    results = []
    added = skipped = failed = 0

    for i, row in enumerate(rows[:300]):
        mapped = {}
        for raw_key, val in row.items():
            key = COL_MAP.get(raw_key.lower().strip(), "")
            if key and str(val).strip():
                mapped[key] = str(val).strip()

        if not mapped.get("Make") and not mapped.get("VIN") and not mapped.get("Internal_ID"):
            continue  # blank row

        # Status normalise
        if mapped.get("Status"):
            mapped["Status"] = STATUS_MAP.get(mapped["Status"].lower(), mapped["Status"].title())
        if not mapped.get("Status"):
            mapped["Status"] = "Available"

        # Dedup
        vin = mapped.get("VIN","").upper()
        vid = mapped.get("Internal_ID","").upper()
        if vin and vin in existing_vins:
            results.append({"row": i+2, "status": "skipped", "msg": "VIN already exists: " + vin,
                            "id": vid or vin[:8]})
            skipped += 1; continue
        if vid and vid in existing_ids:
            results.append({"row": i+2, "status": "skipped", "msg": "Stock # already exists: " + vid,
                            "id": vid})
            skipped += 1; continue

        # Auto-assign ID if missing
        if not mapped.get("Internal_ID"):
            mapped["Internal_ID"] = next_id(existing, "IMP")

        # Cast numeric fields
        if mapped.get("Year"):
            try: mapped["Year"] = int(str(mapped["Year"])[:4])
            except: pass
        for _nf in ("Price","Purchase_Price","Recon_Cost"):
            if mapped.get(_nf):
                mapped[_nf] = _to_num(mapped[_nf]) or mapped[_nf]
        # Mileage stays as string (Airtable text field)

        ok, msg = create_vehicle_record(mapped)
        if ok:
            existing_ids.add(mapped["Internal_ID"].upper())
            if vin: existing_vins.add(vin)
            results.append({"row": i+2, "status": "added", "msg": "Added: " + mapped["Internal_ID"],
                            "id": mapped["Internal_ID"]})
            added += 1
            log_activity("bulk_import", mapped["Internal_ID"], "Bulk import", current_user())
        else:
            results.append({"row": i+2, "status": "failed", "msg": "Error: " + msg[:80],
                            "id": vid or str(i+2)})
            failed += 1

    return jsonify({"ok": True, "added": added, "skipped": skipped, "failed": failed,
                    "total": len(results), "results": results})

@app.errorhandler(404)
def not_found(e): return render_template("404.html"), 404

@app.errorhandler(403)
def forbidden(e): return render_template("403.html"), 403

if __name__ == "__main__":
    app.run(debug=os.environ.get("FLASK_DEBUG","0")=="1", port=5000)


# ══════════════════════════════════════════════════════════════════════════════
# AUTOTRADER EXPANSION — Phase 1: Photo Management
# ══════════════════════════════════════════════════════════════════════════════

import base64, mimetypes
CLOUDINARY_CLOUD = os.environ.get("CLOUDINARY_CLOUD_NAME","")
CLOUDINARY_KEY   = os.environ.get("CLOUDINARY_API_KEY","")
CLOUDINARY_SECRET= os.environ.get("CLOUDINARY_API_SECRET","")
HAS_CLOUDINARY   = all([CLOUDINARY_CLOUD, CLOUDINARY_KEY, CLOUDINARY_SECRET])

def cloudinary_upload(file_bytes, filename, folder="keytrack"):
    """Upload image bytes to Cloudinary, return secure_url or None."""
    if not HAS_CLOUDINARY: return None
    import hashlib, time, hmac
    ts = str(int(time.time()))
    public_id = folder + "/" + filename.rsplit(".",1)[0] + "_" + ts
    sig_str = f"folder={folder}&public_id={public_id}&timestamp={ts}{CLOUDINARY_SECRET}"
    sig = hashlib.sha1(sig_str.encode()).hexdigest()
    try:
        r = requests.post(
            f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD}/image/upload",
            data={"api_key":CLOUDINARY_KEY,"timestamp":ts,"signature":sig,
                  "folder":folder,"public_id":public_id},
            files={"file": (filename, file_bytes)}, timeout=30)
        if r.status_code == 200:
            return r.json().get("secure_url")
    except Exception:
        pass
    return None

def cloudinary_delete(public_id):
    import hashlib, time
    if not HAS_CLOUDINARY: return False
    ts = str(int(time.time()))
    sig_str = f"public_id={public_id}&timestamp={ts}{CLOUDINARY_SECRET}"
    sig = hashlib.sha1(sig_str.encode()).hexdigest()
    try:
        r = requests.post(
            f"https://api.cloudinary.com/v1_1/{CLOUDINARY_CLOUD}/image/destroy",
            data={"api_key":CLOUDINARY_KEY,"timestamp":ts,"signature":sig,"public_id":public_id},
            timeout=15)
        return r.status_code == 200
    except Exception:
        return False

# In-memory photo store: { internal_id: [ {url, public_id, caption, primary} ] }
_vehicle_photos = defaultdict(list)

@app.route("/api/photos/<vid>", methods=["GET"])
@login_required
def api_get_photos(vid):
    return jsonify({"ok":True,"photos":_vehicle_photos.get(vid.upper(),[])})

@app.route("/api/photos/<vid>/upload", methods=["POST"])
@login_required
def api_upload_photo(vid):
    vid = vid.upper()
    files = request.files.getlist("photos")
    if not files:
        return jsonify({"ok":False,"message":"No files uploaded"})
    if len(_vehicle_photos[vid]) + len(files) > 20:
        return jsonify({"ok":False,"message":"Maximum 20 photos per vehicle"})
    added = []
    for f in files:
        if not f.filename: continue
        ext = f.filename.rsplit(".",1)[-1].lower()
        if ext not in ("jpg","jpeg","png","webp","heic"):
            continue
        raw = f.read()
        if len(raw) > 10 * 1024 * 1024:
            continue  # skip >10MB
        url = cloudinary_upload(raw, f"{vid}_{f.filename}", folder="keytrack")
        if not url:
            # Fallback: store as base64 data URL (works without Cloudinary, limited size)
            mime = mimetypes.guess_type(f.filename)[0] or "image/jpeg"
            b64  = base64.b64encode(raw).decode()
            url  = f"data:{mime};base64,{b64}"
        photo = {
            "id": str(uuid.uuid4())[:8],
            "url": url,
            "public_id": f"keytrack/{vid}_{f.filename}",
            "caption": "",
            "primary": len(_vehicle_photos[vid]) == 0,
            "uploaded": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"),
            "staff": current_user(),
        }
        _vehicle_photos[vid].append(photo)
        added.append(photo)
    log_activity("photos_added", vid, f"{len(added)} photo(s) uploaded", current_user())
    return jsonify({"ok":True,"added":len(added),"photos":_vehicle_photos[vid]})

@app.route("/api/photos/<vid>/delete", methods=["POST"])
@login_required
def api_delete_photo(vid):
    vid = vid.upper()
    pid = request.get_json(silent=True).get("id","")
    photos = _vehicle_photos[vid]
    before = len(photos)
    _vehicle_photos[vid] = [p for p in photos if p["id"] != pid]
    # If we deleted the primary, make the first remaining primary
    if _vehicle_photos[vid] and not any(p["primary"] for p in _vehicle_photos[vid]):
        _vehicle_photos[vid][0]["primary"] = True
    return jsonify({"ok": len(_vehicle_photos[vid]) < before})

@app.route("/api/photos/<vid>/set-primary", methods=["POST"])
@login_required
def api_set_primary_photo(vid):
    vid = vid.upper()
    pid = request.get_json(silent=True).get("id","")
    for p in _vehicle_photos[vid]:
        p["primary"] = (p["id"] == pid)
    return jsonify({"ok":True})

@app.route("/api/photos/<vid>/caption", methods=["POST"])
@login_required
def api_caption_photo(vid):
    vid = vid.upper()
    d = request.get_json(silent=True) or {}
    for p in _vehicle_photos[vid]:
        if p["id"] == d.get("id"):
            p["caption"] = d.get("caption","")[:100]
            return jsonify({"ok":True})
    return jsonify({"ok":False})

# ── Photo management page ─────────────────────────────────────────────────────
@app.route("/photos/<vid>")
@login_required
def photos_page(vid):
    vehicle = get_vehicle(vid.upper())
    if not vehicle: abort(404)
    photos = _vehicle_photos.get(vid.upper(), [])
    return render_template("photos.html", vehicle=vehicle,
                           internal_id=vid.upper(), photos=photos,
                           dealership=DEALERSHIP_NAME,
                           has_cloudinary=HAS_CLOUDINARY)


# ══════════════════════════════════════════════════════════════════════════════
# AUTOTRADER EXPANSION — Phase 2: Public Listings
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/listings")
def public_listings():
    """Public buyer-facing listings browse page — no login required."""
    all_v = get_all_vehicles(include_archived=False)
    # Only show Available vehicles publicly
    available = [v for v in all_v if v.get("Status") == "Available"]

    # Filters from query params
    make_filter  = request.args.get("make","").strip()
    body_filter  = request.args.get("body","").strip()
    min_price    = request.args.get("min_price","").strip()
    max_price    = request.args.get("max_price","").strip()
    max_mileage  = request.args.get("max_mileage","").strip()
    sort_by      = request.args.get("sort","newest")
    search_q     = request.args.get("q","").strip().lower()

    filtered = []
    for v in available:
        # Search filter
        if search_q:
            hay = (str(v.get("Make","")) + " " + str(v.get("Model","")) + " " +
                   str(v.get("Year","")) + " " + str(v.get("Trim","")) + " " +
                   str(v.get("Body",""))).lower()
            if search_q not in hay: continue
        # Make filter
        if make_filter and v.get("Make","").lower() != make_filter.lower(): continue
        # Body filter
        if body_filter and body_filter.lower() not in str(v.get("Body","")).lower(): continue
        # Price filters
        try:
            p = float(re.sub(r"[^\d.]","",str(v.get("Price","") or "")) or 0)
            if min_price and p < float(min_price): continue
            if max_price and p > float(max_price): continue
        except: pass
        # Mileage filter
        if max_mileage:
            try:
                m = int(re.sub(r"[^\d]","",str(v.get("Mileage","") or "")) or 0)
                if m > int(max_mileage): continue
            except: pass
        filtered.append(v)

    # Sort
    def price_num(v):
        try: return float(re.sub(r"[^\d.]","",str(v.get("Price","") or "")) or 0)
        except: return 0
    if sort_by == "price_asc":  filtered.sort(key=price_num)
    elif sort_by == "price_desc": filtered.sort(key=price_num, reverse=True)
    elif sort_by == "mileage":
        filtered.sort(key=lambda v: int(re.sub(r"[^\d]","",str(v.get("Mileage","") or "")) or 0))
    elif sort_by == "year": filtered.sort(key=lambda v: int(v.get("Year",0) or 0), reverse=True)
    # Default newest = by Internal_ID descending
    else: filtered.sort(key=lambda v: str(v.get("Internal_ID","")), reverse=True)

    # Build filter options from full available list
    makes = sorted(set(v.get("Make","") for v in available if v.get("Make")))
    bodies = sorted(set(v.get("Body","") for v in available if v.get("Body")))

    # Attach photos
    for v in filtered:
        vid = v.get("Internal_ID","").upper()
        photos = _vehicle_photos.get(vid, [])
        v["_photos"] = photos
        v["_primary_photo"] = next((p["url"] for p in photos if p.get("primary")), None)
        v["_photo_count"] = len(photos)

    return render_template("public_listings.html",
        vehicles=filtered, total=len(filtered),
        makes=makes, bodies=bodies,
        make_filter=make_filter, body_filter=body_filter,
        min_price=min_price, max_price=max_price,
        max_mileage=max_mileage, sort_by=sort_by, search_q=search_q,
        dealership=DEALERSHIP_NAME,
        dealership_settings=_settings)

@app.route("/listing/<internal_id>")
def public_listing_detail(internal_id):
    """Full public listing page for a single vehicle — SEO optimised."""
    vehicle = get_vehicle(internal_id.upper())
    if not vehicle or vehicle.get("Status") == "Archived": abort(404)
    photos  = _vehicle_photos.get(internal_id.upper(), [])
    primary = next((p["url"] for p in photos if p.get("primary")), None)
    finance = gen_finance_options(vehicle, deposit=0, term_years=5)
    # Related vehicles (same make, different ID)
    all_v = get_all_vehicles()
    related = [v for v in all_v
               if v.get("Make","").lower() == vehicle.get("Make","").lower()
               and v.get("Internal_ID","") != internal_id.upper()
               and v.get("Status") == "Available"][:3]
    for rv in related:
        rid = rv.get("Internal_ID","").upper()
        rv["_primary_photo"] = next((p["url"] for p in _vehicle_photos.get(rid,[]) if p.get("primary")), None)
    log_activity("public_view", internal_id.upper(), "Public listing viewed", "public")
    # Pass all contact info so template can render Call Now panel
    contact_phone   = _settings.get("contact_phone", DEALERSHIP_PHONE or MANAGER_PHONE)
    contact_address = _settings.get("address", DEALERSHIP_ADDRESS)
    return render_template("public_listing_detail.html",
        vehicle=vehicle, internal_id=internal_id.upper(),
        photos=photos, primary_photo=primary,
        finance=finance, related=related,
        dealership=DEALERSHIP_NAME,
        dealership_settings=_settings,
        manager_phone=contact_phone,
        contact_phone=contact_phone,
        contact_address=contact_address)

@app.route("/api/public-enquiry", methods=["POST"])
def api_public_enquiry():
    """Public enquiry form submission — no login needed."""
    d = request.get_json(silent=True) or {}
    name  = d.get("name","").strip()
    phone = d.get("phone","").strip()
    vid   = d.get("internal_id","").upper()
    if not name or not phone:
        return jsonify({"ok":False,"message":"Name and phone are required"})
    enquiry = {
        "id": str(uuid.uuid4())[:8],
        "internal_id": vid,
        "name": name,
        "phone": phone,
        "email": d.get("email","").strip(),
        "notes": d.get("message","").strip(),
        "source": "website",
        "follow_up_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "follow_up_status": "pending",
        "timestamp": datetime.now(timezone.utc).strftime("%b %d %H:%M"),
        "timestamp_iso": datetime.now(timezone.utc).isoformat(),
        "staff": "website",
    }
    _enquiries.insert(0, enquiry)
    log_activity("web_enquiry", vid, f"{name} — {phone}", "website")
    # SMS notification to manager
    vehicle = get_vehicle(vid) if vid else {}
    vname = f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')}".strip()
    sms = f"NEW ENQUIRY: {name} ({phone}) interested in {vname or vid}. Check KeyTrack."
    send_sms(MANAGER_PHONE, sms)
    return jsonify({"ok":True,"message":"Thanks! We'll be in touch shortly."})

@app.route("/api/public-test-drive", methods=["POST"])
def api_public_test_drive():
    """Public test drive booking — no login needed."""
    d = request.get_json(silent=True) or {}
    customer = d.get("name","").strip()
    vid = d.get("internal_id","").upper()
    dt  = d.get("datetime","").strip()
    if not all([customer, vid, dt]):
        return jsonify({"ok":False,"message":"Name, vehicle and preferred time required"})
    booking = {
        "id": str(uuid.uuid4())[:8], "internal_id": vid,
        "datetime": dt, "customer_name": customer,
        "phone": d.get("phone","").strip(),
        "notes": "Booked via website",
        "status": "scheduled", "staff": "website",
        "created": datetime.now(timezone.utc).isoformat(),
    }
    _test_drives.insert(0, booking)
    log_activity("web_test_drive", vid, f"{customer} @ {dt}", "website")
    vehicle = get_vehicle(vid) if vid else {}
    vname = f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')}".strip()
    send_sms(MANAGER_PHONE, f"TEST DRIVE BOOKED: {customer} for {vname} on {dt}")
    return jsonify({"ok":True,"message":"Test drive request received! We'll confirm shortly."})


# ══════════════════════════════════════════════════════════════════════════════
# AUTOTRADER EXPANSION — Phase 3: Marketplace Feed Exports
# ══════════════════════════════════════════════════════════════════════════════

def _vehicle_price_num(v):
    try: return int(float(re.sub(r"[^\d.]","",str(v.get("Price","") or "")) or 0))
    except: return 0

def _vehicle_mileage_num(v):
    try: return int(re.sub(r"[^\d]","",str(v.get("Mileage","") or "")) or 0)
    except: return 0

@app.route("/feed/carsales.xml")
def feed_carsales():
    """Carsales.com.au compatible XML feed."""
    all_v = get_all_vehicles(include_archived=False)
    available = [v for v in all_v if v.get("Status") == "Available"]
    base_url = request.host_url.rstrip("/")
    lines = ['<?xml version="1.0" encoding="UTF-8"?>']
    lines.append('<Inventory xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">')
    lines.append(f'  <Provider>{DEALERSHIP_NAME}</Provider>')
    lines.append(f'  <LastUpdated>{datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")}</LastUpdated>')
    for v in available:
        vid = v.get("Internal_ID","")
        photos = _vehicle_photos.get(vid.upper(), [])
        photo_urls = [p["url"] for p in photos if not p["url"].startswith("data:")]
        lines.append('  <Vehicle>')
        lines.append(f'    <StockNumber>{vid}</StockNumber>')
        lines.append(f'    <VIN>{v.get("VIN","")}</VIN>')
        lines.append(f'    <Year>{v.get("Year","")}</Year>')
        lines.append(f'    <Make>{v.get("Make","")}</Make>')
        lines.append(f'    <Model>{v.get("Model","")}</Model>')
        lines.append(f'    <Variant>{v.get("Trim","")}</Variant>')
        lines.append(f'    <BodyStyle>{v.get("Body","")}</BodyStyle>')
        lines.append(f'    <Transmission>{v.get("Transmission","")}</Transmission>')
        lines.append(f'    <Drive>{v.get("Drive","")}</Drive>')
        lines.append(f'    <FuelType>{v.get("Fuel","")}</FuelType>')
        lines.append(f'    <Odometer unit="km">{_vehicle_mileage_num(v)}</Odometer>')
        lines.append(f'    <Colour>{v.get("Color","")}</Colour>')
        lines.append(f'    <Engine>{v.get("Engine","")}</Engine>')
        lines.append(f'    <Price>{_vehicle_price_num(v)}</Price>')
        lines.append(f'    <Comments><![CDATA[{v.get("Notes","")}]]></Comments>')
        lines.append(f'    <ListingURL>{base_url}/listing/{vid}</ListingURL>')
        lines.append('    <Photos>')
        for url in photo_urls[:10]:
            lines.append(f'      <Photo>{url}</Photo>')
        lines.append('    </Photos>')
        lines.append('  </Vehicle>')
    lines.append('</Inventory>')
    xml = "\n".join(lines)
    resp = make_response(xml)
    resp.headers["Content-Type"] = "application/xml; charset=utf-8"
    resp.headers["Cache-Control"] = "public, max-age=900"
    return resp

@app.route("/feed/facebook.json")
def feed_facebook():
    """Facebook / Instagram Catalogue JSON feed."""
    all_v = get_all_vehicles(include_archived=False)
    available = [v for v in all_v if v.get("Status") == "Available"]
    base_url = request.host_url.rstrip("/")
    items = []
    for v in available:
        vid = v.get("Internal_ID","")
        photos = _vehicle_photos.get(vid.upper(), [])
        primary = next((p["url"] for p in photos if p.get("primary") and not p["url"].startswith("data:")), "")
        price = _vehicle_price_num(v)
        name = f"{v.get('Year','')} {v.get('Make','')} {v.get('Model','')} {v.get('Trim','')}".strip()
        desc = v.get("Notes","") or f"{v.get('Year','')} {v.get('Make','')} {v.get('Model','')} — {v.get('Engine','')} — {v.get('Mileage','')}km"
        items.append({
            "id": vid,
            "title": name,
            "description": desc[:500],
            "availability": "in stock",
            "condition": "used",
            "price": f"{price} AUD",
            "link": f"{base_url}/listing/{vid}",
            "image_link": primary,
            "brand": v.get("Make",""),
            "vehicle_model_year": str(v.get("Year","")),
            "make": v.get("Make",""),
            "model": v.get("Model",""),
            "mileage": {"value": _vehicle_mileage_num(v), "unit": "KM"},
            "body_style": v.get("Body",""),
            "transmission": v.get("Transmission",""),
            "fuel_type": v.get("Fuel",""),
            "exterior_color": v.get("Color",""),
            "vin": v.get("VIN",""),
            "stock_number": vid,
        })
    resp = make_response(json.dumps({"data": items}, indent=2))
    resp.headers["Content-Type"] = "application/json; charset=utf-8"
    resp.headers["Cache-Control"] = "public, max-age=900"
    return resp

@app.route("/feed/autotrader.xml")
def feed_autotrader():
    """AutoTrader AU compatible XML feed (STARX format)."""
    all_v = get_all_vehicles(include_archived=False)
    available = [v for v in all_v if v.get("Status") == "Available"]
    base_url = request.host_url.rstrip("/")
    lines = ['<?xml version="1.0" encoding="UTF-8"?>','<stockxml>','  <dealer>',
             f'    <name>{DEALERSHIP_NAME}</name>',
             f'    <updated>{datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")}</updated>',
             '  </dealer>','  <stock>']
    for v in available:
        vid = v.get("Internal_ID","")
        photos = _vehicle_photos.get(vid.upper(), [])
        photo_urls = [p["url"] for p in photos if not p["url"].startswith("data:")]
        lines += ['    <vehicle>',
            f'      <unique_id>{vid}</unique_id>',
            f'      <rego_plate></rego_plate>',
            f'      <vin>{v.get("VIN","")}</vin>',
            f'      <year>{v.get("Year","")}</year>',
            f'      <make>{v.get("Make","")}</make>',
            f'      <family>{v.get("Model","")}</family>',
            f'      <variant>{v.get("Trim","")}</variant>',
            f'      <series></series>',
            f'      <body_type>{v.get("Body","")}</body_type>',
            f'      <transmission>{v.get("Transmission","")}</transmission>',
            f'      <drive_type>{v.get("Drive","")}</drive_type>',
            f'      <fuel_type>{v.get("Fuel","")}</fuel_type>',
            f'      <engine>{v.get("Engine","")}</engine>',
            f'      <odometer>{_vehicle_mileage_num(v)}</odometer>',
            f'      <colour>{v.get("Color","")}</colour>',
            f'      <price>{_vehicle_price_num(v)}</price>',
            f'      <comments><![CDATA[{v.get("Notes","")}]]></comments>',
            f'      <url>{base_url}/listing/{vid}</url>',
            '      <images>']
        for url in photo_urls[:15]:
            lines.append(f'        <image>{url}</image>')
        lines += ['      </images>','    </vehicle>']
    lines += ['  </stock>','</stockxml>']
    resp = make_response("\n".join(lines))
    resp.headers["Content-Type"] = "application/xml; charset=utf-8"
    resp.headers["Cache-Control"] = "public, max-age=900"
    return resp

@app.route("/feed/gumtree.csv")
def feed_gumtree():
    """Gumtree Motors CSV export."""
    all_v = get_all_vehicles(include_archived=False)
    available = [v for v in all_v if v.get("Status") == "Available"]
    base_url = request.host_url.rstrip("/")
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Title","Price","Description","Year","Make","Model","Body Type",
                     "Transmission","Kilometres","Fuel Type","Colour","VIN","Stock#","URL","Photo 1","Photo 2","Photo 3"])
    for v in available:
        vid = v.get("Internal_ID","")
        photos = [p["url"] for p in _vehicle_photos.get(vid.upper(),[]) if not p["url"].startswith("data:")]
        title = f"{v.get('Year','')} {v.get('Make','')} {v.get('Model','')} {v.get('Trim','')}".strip()
        writer.writerow([
            title, _vehicle_price_num(v),
            (v.get("Notes","") or title)[:500],
            v.get("Year",""), v.get("Make",""), v.get("Model",""),
            v.get("Body",""), v.get("Transmission",""),
            _vehicle_mileage_num(v), v.get("Fuel",""), v.get("Color",""),
            v.get("VIN",""), vid,
            f"{base_url}/listing/{vid}",
            photos[0] if len(photos)>0 else "",
            photos[1] if len(photos)>1 else "",
            photos[2] if len(photos)>2 else "",
        ])
    resp = make_response(output.getvalue())
    resp.headers["Content-Type"] = "text/csv"
    resp.headers["Content-Disposition"] = f'attachment; filename=gumtree-{datetime.now().strftime("%Y%m%d")}.csv'
    return resp

# ── Feed dashboard ────────────────────────────────────────────────────────────
@app.route("/admin/feeds")
@admin_required
def feeds_page():
    all_v = get_all_vehicles(include_archived=False)
    available = [v for v in all_v if v.get("Status") == "Available"]
    base_url = request.host_url.rstrip("/")
    return render_template("feeds.html",
        dealership=DEALERSHIP_NAME,
        vehicle_count=len(available),
        base_url=base_url,
        has_cloudinary=HAS_CLOUDINARY)


# ══════════════════════════════════════════════════════════════════════════════
# AUTOTRADER EXPANSION — Phase 4: Sold Scoreboard & Price Drop Badges
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/scoreboard")
@login_required
def scoreboard_page():
    """TV wall display — daily/weekly/monthly sales scoreboard."""
    all_v = get_all_vehicles(include_archived=True)
    now   = datetime.now()
    sold  = [v for v in all_v if v.get("Status") == "Sold"]

    def in_period(v, days):
        sd = v.get("Sold_Date","")
        if not sd: return False
        try:
            d = datetime.fromisoformat(sd[:10])
            return (now - d).days <= days
        except: return False

    today_sales   = [v for v in sold if in_period(v, 0)]
    week_sales    = [v for v in sold if in_period(v, 7)]
    month_sales   = [v for v in sold if in_period(v, 30)]
    target        = _settings.get("monthly_target", 15)
    pct           = min(100, round(len(month_sales) / target * 100)) if target else 0

    # Revenue
    def total_rev(vehicles):
        t = 0
        for v in vehicles:
            try: t += float(re.sub(r"[^\d.]","",str(v.get("Price","") or "")) or 0)
            except: pass
        return round(t)

    return render_template("scoreboard.html",
        dealership=DEALERSHIP_NAME,
        today_sales=today_sales, today_count=len(today_sales),
        week_sales=week_sales,   week_count=len(week_sales),
        month_sales=month_sales, month_count=len(month_sales),
        target=target, pct=pct,
        today_revenue=total_rev(today_sales),
        week_revenue=total_rev(week_sales),
        month_revenue=total_rev(month_sales),
        all_vehicles=all_v,
        stock_value=total_rev([v for v in all_v if v.get("Status")=="Available"]),
    )

# ── Price drop badge API ──────────────────────────────────────────────────────
@app.route("/api/price-drops")
@login_required
def api_price_drops():
    """Return vehicles whose price was recently reduced."""
    drops = []
    for vid, history in _price_history.items():
        if len(history) < 2: continue
        latest = history[-1]["price"]
        previous = history[-2]["price"]
        try:
            curr = float(re.sub(r"[^\d.]","",str(latest or "")) or 0)
            prev = float(re.sub(r"[^\d.]","",str(previous or "")) or 0)
            if curr < prev and prev > 0:
                drop_amt = round(prev - curr)
                drop_pct = round((prev - curr) / prev * 100)
                drops.append({
                    "vid": vid, "from": prev, "to": curr,
                    "drop_amount": drop_amt, "drop_pct": drop_pct,
                    "date": history[-1]["date"],
                })
        except: pass
    return jsonify({"ok":True,"drops":drops})



# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 1 — BUYER ACCOUNTS + SAVED VEHICLES + PRICE DROP ALERTS
# ══════════════════════════════════════════════════════════════════════════════


# In-memory buyer store  { email: { password_hash, name, phone, created, saved:[], alerts:[], sessions:[] } }
_buyers = {}
# In-memory buyer sessions { token: email }
_buyer_sessions = {}

def _buyer_token():
    return secrets.token_hex(32)

def _current_buyer():
    t = request.cookies.get("buyer_token","")
    return _buyer_sessions.get(t)

def buyer_login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*a,**kw):
        if not _current_buyer():
            return redirect("/buyer/login?next="+request.path)
        return f(*a,**kw)
    return decorated

# ── Registration / Login ──────────────────────────────────────────────────────
@app.route("/buyer/register", methods=["GET","POST"])
def buyer_register():
    if request.method == "GET":
        return render_template("buyer_register.html", dealership=DEALERSHIP_NAME)
    d = request.get_json(silent=True) or request.form
    email = str(d.get("email","")).strip().lower()
    name  = str(d.get("name","")).strip()
    phone = str(d.get("phone","")).strip()
    pw    = str(d.get("password",""))
    if not email or not pw or not name:
        return jsonify({"ok":False,"message":"Name, email and password required"})
    if email in _buyers:
        return jsonify({"ok":False,"message":"Email already registered"})
    if len(pw) < 6:
        return jsonify({"ok":False,"message":"Password must be at least 6 characters"})
    _buyers[email] = {
        "email": email, "name": name, "phone": phone,
        "password_hash": generate_password_hash(pw),
        "created": datetime.now(timezone.utc).isoformat(),
        "saved": [],        # list of internal_ids
        "alerts": [],       # list of { type, value, label }
        "notifications": [], # list of { message, vid, date, read }
    }
    token = _buyer_token()
    _buyer_sessions[token] = email
    resp = jsonify({"ok":True,"redirect":"/buyer/dashboard"})
    resp.set_cookie("buyer_token", token, max_age=86400*30, httponly=True, samesite="Lax")
    log_activity("buyer_register", email, f"{name} registered", "website")
    return resp

@app.route("/buyer/login", methods=["GET","POST"])
def buyer_login():
    if request.method == "GET":
        return render_template("buyer_login.html", dealership=DEALERSHIP_NAME,
                               next=request.args.get("next","/buyer/dashboard"))
    d = request.get_json(silent=True) or request.form
    email = str(d.get("email","")).strip().lower()
    pw    = str(d.get("password",""))
    buyer = _buyers.get(email)
    if not buyer or not check_password_hash(buyer["password_hash"], pw):
        return jsonify({"ok":False,"message":"Invalid email or password"})
    token = _buyer_token()
    _buyer_sessions[token] = email
    resp = jsonify({"ok":True,"redirect": str(d.get("next","/buyer/dashboard"))})
    resp.set_cookie("buyer_token", token, max_age=86400*30, httponly=True, samesite="Lax")
    return resp

@app.route("/buyer/logout")
def buyer_logout():
    t = request.cookies.get("buyer_token","")
    _buyer_sessions.pop(t, None)
    resp = redirect("/listings")
    resp.delete_cookie("buyer_token")
    return resp

# ── Buyer dashboard ───────────────────────────────────────────────────────────
@app.route("/buyer/dashboard")
@buyer_login_required
def buyer_dashboard():
    email = _current_buyer()
    buyer = _buyers[email]
    saved_vehicles = []
    for vid in buyer["saved"]:
        v = get_vehicle(vid)
        if v:
            v["_primary_photo"] = next((p["url"] for p in _vehicle_photos.get(vid,[]) if p.get("primary")), None)
            saved_vehicles.append(v)
    unread = sum(1 for n in buyer.get("notifications",[]) if not n.get("read"))
    return render_template("buyer_dashboard.html",
        buyer=buyer, saved_vehicles=saved_vehicles,
        dealership=DEALERSHIP_NAME, unread=unread)

# ── Save / unsave vehicle ─────────────────────────────────────────────────────
@app.route("/api/buyer/save", methods=["POST"])
def api_buyer_save():
    email = _current_buyer()
    if not email:
        return jsonify({"ok":False,"login_required":True})
    d = request.get_json(silent=True) or {}
    vid = d.get("vid","").upper()
    buyer = _buyers[email]
    if vid in buyer["saved"]:
        buyer["saved"].remove(vid)
        return jsonify({"ok":True,"saved":False,"count":len(buyer["saved"])})
    else:
        buyer["saved"].append(vid)
        return jsonify({"ok":True,"saved":True,"count":len(buyer["saved"])})

@app.route("/api/buyer/saved")
def api_buyer_saved():
    email = _current_buyer()
    if not email: return jsonify({"ok":True,"saved":[]})
    return jsonify({"ok":True,"saved":_buyers[email]["saved"]})

# ── Search alerts ─────────────────────────────────────────────────────────────
@app.route("/api/buyer/alerts", methods=["GET"])
@buyer_login_required
def api_buyer_get_alerts():
    buyer = _buyers[_current_buyer()]
    return jsonify({"ok":True,"alerts":buyer["alerts"]})

@app.route("/api/buyer/alerts/add", methods=["POST"])
@buyer_login_required
def api_buyer_add_alert():
    buyer = _buyers[_current_buyer()]
    d = request.get_json(silent=True) or {}
    alert = {
        "id": str(uuid.uuid4())[:8],
        "make": d.get("make","").strip(),
        "model": d.get("model","").strip(),
        "max_price": d.get("max_price",""),
        "body": d.get("body","").strip(),
        "label": d.get("label","New vehicle alert"),
        "created": datetime.now(timezone.utc).strftime("%b %d %Y"),
        "active": True,
    }
    buyer["alerts"].append(alert)
    return jsonify({"ok":True,"alert":alert})

@app.route("/api/buyer/alerts/delete", methods=["POST"])
@buyer_login_required
def api_buyer_delete_alert():
    buyer = _buyers[_current_buyer()]
    aid = (request.get_json(silent=True) or {}).get("id","")
    buyer["alerts"] = [a for a in buyer["alerts"] if a["id"] != aid]
    return jsonify({"ok":True})

@app.route("/api/buyer/notifications")
@buyer_login_required
def api_buyer_notifications():
    buyer = _buyers[_current_buyer()]
    # Mark all read
    for n in buyer.get("notifications",[]): n["read"] = True
    return jsonify({"ok":True,"notifications":buyer.get("notifications",[])})

# ── Price drop check (run whenever a vehicle price changes) ──────────────────
def _notify_buyers_price_drop(vid, old_price, new_price):
    """Called when a vehicle price drops — notify buyers who saved it."""
    drop_amt = old_price - new_price
    if drop_amt <= 0: return
    vehicle = get_vehicle(vid)
    if not vehicle: return
    vname = f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')}".strip()
    for email, buyer in _buyers.items():
        if vid in buyer.get("saved",[]):
            note = {
                "id": str(uuid.uuid4())[:8],
                "type": "price_drop",
                "message": f"Price drop on {vname} — saved ${drop_amt:,.0f}! Now ${new_price:,.0f}",
                "vid": vid, "vname": vname,
                "old_price": old_price, "new_price": new_price,
                "date": datetime.now(timezone.utc).strftime("%b %d %H:%M"),
                "read": False,
            }
            buyer.setdefault("notifications",[]).insert(0, note)
            # SMS if they have a phone
            if buyer.get("phone"):
                try:
                    from flask import has_request_context
                    base = request.host_url if has_request_context() else ""
                except Exception:
                    base = ""
                send_sms(buyer["phone"],
                    f"Price drop! {vname} now ${new_price:,.0f} (was ${old_price:,.0f}). View: {base}listing/{vid}")

def _notify_buyers_new_listing(vehicle):
    """Called when a new vehicle is added — notify buyers with matching alerts."""
    for email, buyer in _buyers.items():
        for alert in buyer.get("alerts",[]):
            if not alert.get("active"): continue
            make_match  = not alert.get("make")  or alert["make"].lower()  == str(vehicle.get("Make","")).lower()
            model_match = not alert.get("model") or alert["model"].lower() in str(vehicle.get("Model","")).lower()
            body_match  = not alert.get("body")  or alert["body"].lower()  in str(vehicle.get("Body","")).lower()
            price_match = True
            if alert.get("max_price"):
                try:
                    price_match = _vehicle_price_num(vehicle) <= float(alert["max_price"])
                except: pass
            if make_match and model_match and body_match and price_match:
                vid = vehicle.get("Internal_ID","")
                vname = f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')}".strip()
                note = {
                    "id": str(uuid.uuid4())[:8],
                    "type": "new_listing",
                    "message": f"New listing matching your alert: {vname} — ${_vehicle_price_num(vehicle):,.0f}",
                    "vid": vid, "vname": vname,
                    "date": datetime.now(timezone.utc).strftime("%b %d %H:%M"),
                    "read": False,
                }
                buyer.setdefault("notifications",[]).insert(0, note)
                if buyer.get("phone"):
                    try:
                        from flask import has_request_context
                        base = request.host_url if has_request_context() else ""
                    except Exception:
                        base = ""
                    send_sms(buyer["phone"],
                        f"New listing alert! {vname} ${_vehicle_price_num(vehicle):,.0f}. View: {base}listing/{vid}")


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 2 — AI PLAIN-ENGLISH SEARCH
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/search-ai")
def ai_search_page():
    buyer = _current_buyer()
    buyer_obj = _buyers.get(buyer) if buyer else None
    return render_template("ai_search.html",
        dealership=DEALERSHIP_NAME, buyer=buyer_obj)

@app.route("/api/ai-search", methods=["POST"])
def api_ai_search():
    """Parse a plain-English query with GPT and return matching vehicles."""
    d = request.get_json(silent=True) or {}
    query = d.get("query","").strip()
    if not query:
        return jsonify({"ok":False,"message":"Enter a search query"})

    all_v = get_all_vehicles(include_archived=False)
    available = [v for v in all_v if v.get("Status") == "Available"]

    # Build context for GPT
    makes  = sorted(set(v.get("Make","")  for v in available if v.get("Make")))
    bodies = sorted(set(v.get("Body","")  for v in available if v.get("Body")))
    fuels  = sorted(set(v.get("Fuel","")  for v in available if v.get("Fuel")))

    system = """You are a vehicle search assistant. Parse the user's plain-English query and extract search filters as JSON.
Return ONLY valid JSON with these optional fields:
{
  "make": "Toyota" or null,
  "model": "Hilux" or null,
  "body": "Ute" or null,
  "fuel": "Diesel" or null,
  "max_price": 30000 or null,
  "min_price": null,
  "max_mileage": 150000 or null,
  "min_year": 2015 or null,
  "transmission": "Automatic" or null,
  "keywords": ["tow", "family", "4WD"] or [],
  "summary": "Short human-readable summary of what you understood"
}
Be flexible — "cheap" means under $15000, "new" means under 50000km, "family" means wagon/SUV/people mover,
"tradie" means ute/van, "fuel efficient" means hybrid/electric/small engine.
Only include fields that are clearly implied. Return null for anything not mentioned."""

    filters = {}
    ai_summary = "Showing all available vehicles"
    try:
        if OPENAI_API_KEY:
            import openai
            client = openai.OpenAI(api_key=OPENAI_API_KEY)
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role":"system","content":system},
                    {"role":"user","content":f"Available makes: {', '.join(makes)}\nAvailable bodies: {', '.join(bodies)}\nQuery: {query}"}
                ],
                max_tokens=400, temperature=0.1
            )
            raw = resp.choices[0].message.content.strip()
            raw = re.sub(r"```json|```","",raw).strip()
            filters = json.loads(raw)
            ai_summary = filters.pop("summary", ai_summary)
    except Exception as e:
        # Fallback: simple keyword matching
        ql = query.lower()
        for make in makes:
            if make.lower() in ql:
                filters["make"] = make; break
        for body in bodies:
            if body.lower() in ql:
                filters["body"] = body; break
        price_match = re.search(r"under\s*\$?([\d,]+)k?", ql)
        if price_match:
            p = float(price_match.group(1).replace(",",""))
            filters["max_price"] = p * 1000 if p < 1000 else p
        ai_summary = f"Searching for: {query}"

    # Apply filters to vehicles
    results = []
    keywords = filters.pop("keywords",[]) or []
    for v in available:
        if filters.get("make") and filters["make"].lower() != str(v.get("Make","")).lower(): continue
        if filters.get("model") and filters["model"].lower() not in str(v.get("Model","")).lower(): continue
        if filters.get("body") and filters["body"].lower() not in str(v.get("Body","")).lower(): continue
        if filters.get("fuel") and filters["fuel"].lower() not in str(v.get("Fuel","")).lower(): continue
        if filters.get("transmission") and filters["transmission"].lower() not in str(v.get("Transmission","")).lower(): continue
        if filters.get("max_price"):
            try:
                if _vehicle_price_num(v) > float(filters["max_price"]): continue
            except: pass
        if filters.get("min_price"):
            try:
                if _vehicle_price_num(v) < float(filters["min_price"]): continue
            except: pass
        if filters.get("max_mileage"):
            try:
                if _vehicle_mileage_num(v) > float(filters["max_mileage"]): continue
            except: pass
        if filters.get("min_year"):
            try:
                if int(v.get("Year",0) or 0) < int(filters["min_year"]): continue
            except: pass
        # Keyword scoring
        score = 0
        haystack = " ".join([str(v.get(f,"")) for f in ["Make","Model","Body","Engine","Fuel","Transmission","Notes","Trim"]]).lower()
        for kw in keywords:
            if kw.lower() in haystack: score += 1
        v["_score"] = score
        v["_primary_photo"] = next((p["url"] for p in _vehicle_photos.get(str(v.get("Internal_ID","")).upper(),[]) if p.get("primary")), None)
        results.append(v)

    # Sort by score then price
    results.sort(key=lambda v: (-v.get("_score",0), _vehicle_price_num(v)))

    # Slim down for JSON response
    slim = []
    for v in results[:20]:
        slim.append({
            "internal_id": v.get("Internal_ID",""),
            "year": v.get("Year",""), "make": v.get("Make",""),
            "model": v.get("Model",""), "trim": v.get("Trim",""),
            "price": v.get("Price",""), "mileage": v.get("Mileage",""),
            "body": v.get("Body",""), "engine": v.get("Engine",""),
            "transmission": v.get("Transmission",""), "color": v.get("Color",""),
            "fuel": v.get("Fuel",""), "status": v.get("Status",""),
            "photo": v.get("_primary_photo",""),
            "score": v.get("_score",0),
        })

    log_activity("ai_search", "public", query, _current_buyer() or "public")
    return jsonify({"ok":True,"results":slim,"total":len(slim),
                    "summary":ai_summary,"filters":filters,"query":query})


# ══════════════════════════════════════════════════════════════════════════════
# FEATURE 3 — MULTI-DEALER SUPPORT
# ══════════════════════════════════════════════════════════════════════════════

# In-memory dealer store { dealer_id: { name, email, phone, address, suburb, state,
#   plan, status, password_hash, airtable_token, airtable_base, logo_url,
#   created, vehicles_count, subscription_end } }
_dealers = {}
# Map airtable_base → dealer_id for fast lookup
_dealer_by_base = {}

PLATFORM_ADMIN_EMAIL    = os.environ.get("PLATFORM_ADMIN_EMAIL","platform@keytrack.com.au")
PLATFORM_ADMIN_PASSWORD = os.environ.get("PLATFORM_ADMIN_PASSWORD","PlatformAdmin#2025!")
PLATFORM_ADMIN_TOKEN    = None

PLANS = {
    "free":       {"name":"Free",       "price":0,   "max_listings":5,  "features":["Basic listings","Feed exports"]},
    "dealer":     {"name":"Dealer",     "price":99,  "max_listings":50, "features":["Unlimited listings","Full CRM","Key tracker","Documents","Scoreboard"]},
    "pro":        {"name":"Pro",        "price":199, "max_listings":200,"features":["Everything in Dealer","AI descriptions","Price intelligence","Buyer alerts","Priority placement"]},
    "enterprise": {"name":"Enterprise", "price":399, "max_listings":999,"features":["Everything in Pro","White label","Custom domain","Dedicated support"]},
}

def _dealer_id():
    return "DLR" + str(uuid.uuid4())[:6].upper()

# ── Platform admin ────────────────────────────────────────────────────────────
@app.route("/platform/login", methods=["GET","POST"])
def platform_login():
    if request.method == "GET":
        return render_template("platform_login.html")
    d = request.get_json(silent=True) or request.form
    if d.get("email") == PLATFORM_ADMIN_EMAIL and d.get("password") == PLATFORM_ADMIN_PASSWORD:
        token = secrets.token_hex(32)
        global PLATFORM_ADMIN_TOKEN
        PLATFORM_ADMIN_TOKEN = token
        resp = jsonify({"ok":True,"redirect":"/platform/admin"})
        resp.set_cookie("platform_token", token, max_age=3600*8, httponly=True, samesite="Lax")
        return resp
    return jsonify({"ok":False,"message":"Invalid credentials"})

def platform_admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*a,**kw):
        t = request.cookies.get("platform_token","")
        if not PLATFORM_ADMIN_TOKEN or t != PLATFORM_ADMIN_TOKEN:
            return redirect("/platform/login")
        return f(*a,**kw)
    return decorated

@app.route("/platform/admin")
@platform_admin_required
def platform_admin():
    dealers = list(_dealers.values())
    total_vehicles = sum(d.get("vehicles_count",0) for d in dealers)
    revenue = sum(PLANS.get(d.get("plan","free"),{}).get("price",0) for d in dealers if d.get("status")=="active")
    return render_template("platform_admin.html",
        dealers=dealers, plans=PLANS,
        total_dealers=len(dealers),
        active_dealers=sum(1 for d in dealers if d.get("status")=="active"),
        total_vehicles=total_vehicles,
        monthly_revenue=revenue)

# ── Dealer onboarding ─────────────────────────────────────────────────────────
@app.route("/dealers/join", methods=["GET"])
def dealer_join():
    return render_template("dealer_join.html", plans=PLANS, dealership=DEALERSHIP_NAME)

@app.route("/api/dealers/register", methods=["POST"])
def api_dealer_register():
    d = request.get_json(silent=True) or {}
    name  = d.get("name","").strip()
    email = d.get("email","").strip().lower()
    phone = d.get("phone","").strip()
    pw    = d.get("password","")
    plan  = d.get("plan","free")
    if not all([name, email, pw]):
        return jsonify({"ok":False,"message":"Name, email and password required"})
    if any(dd.get("email")==email for dd in _dealers.values()):
        return jsonify({"ok":False,"message":"Email already registered"})
    if plan not in PLANS:
        plan = "free"
    did = _dealer_id()
    _dealers[did] = {
        "id": did, "name": name, "email": email, "phone": phone,
        "address": d.get("address","").strip(),
        "suburb": d.get("suburb","").strip(),
        "state": d.get("state","NSW"),
        "plan": plan,
        "status": "pending",  # pending → active after admin approval or payment
        "password_hash": generate_password_hash(pw),
        "airtable_token": d.get("airtable_token","").strip(),
        "airtable_base": d.get("airtable_base","").strip(),
        "logo_url": "",
        "created": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "vehicles_count": 0,
        "subscription_end": "",
        "slug": re.sub(r"[^a-z0-9]+","-",name.lower()).strip("-"),
    }
    if d.get("airtable_base"):
        _dealer_by_base[d["airtable_base"].strip()] = did
    log_activity("dealer_register", did, f"{name} — {plan} plan", "website")
    # Notify platform admin
    send_sms(MANAGER_PHONE, f"NEW DEALER: {name} ({email}) signed up for {plan} plan. Review at /platform/admin")
    return jsonify({"ok":True,"dealer_id":did,
                    "message":"Registration received! We'll activate your account within 24 hours."})

@app.route("/api/platform/dealers/<did>/activate", methods=["POST"])
@platform_admin_required
def platform_activate_dealer(did):
    if did not in _dealers:
        return jsonify({"ok":False,"message":"Dealer not found"})
    _dealers[did]["status"] = "active"
    log_activity("dealer_activated", did, _dealers[did]["name"], "platform")
    return jsonify({"ok":True})

@app.route("/api/platform/dealers/<did>/plan", methods=["POST"])
@platform_admin_required
def platform_change_plan(did):
    if did not in _dealers: return jsonify({"ok":False})
    plan = (request.get_json(silent=True) or {}).get("plan","free")
    if plan in PLANS:
        _dealers[did]["plan"] = plan
    return jsonify({"ok":True})

@app.route("/api/platform/dealers/<did>/delete", methods=["POST"])
@platform_admin_required
def platform_delete_dealer(did):
    _dealers.pop(did, None)
    return jsonify({"ok":True})

# ── Public dealer directory ───────────────────────────────────────────────────
@app.route("/dealers")
def dealer_directory():
    active = [d for d in _dealers.values() if d.get("status")=="active"]
    return render_template("dealer_directory.html",
        dealers=active, dealership=DEALERSHIP_NAME, plans=PLANS)

@app.route("/dealer/<slug>")
def dealer_profile(slug):
    dealer = next((d for d in _dealers.values() if d.get("slug")==slug and d.get("status")=="active"), None)
    if not dealer: abort(404)
    # Get their vehicles (from global pool filtered by dealer's airtable base, or tag)
    # For now show all vehicles tagged to this dealer
    all_v = get_all_vehicles(include_archived=False)
    dealer_vehicles = [v for v in all_v
                       if v.get("Status")=="Available"
                       and str(v.get("Dealer_ID","")) == dealer["id"]]
    for v in dealer_vehicles:
        vid = v.get("Internal_ID","").upper()
        v["_primary_photo"] = next((p["url"] for p in _vehicle_photos.get(vid,[]) if p.get("primary")), None)
    return render_template("dealer_profile.html",
        dealer=dealer, vehicles=dealer_vehicles,
        dealership=DEALERSHIP_NAME, plans=PLANS)



# ══════════════════════════════════════════════════════════════════════════════
# PREMIUM EXPANSION — AUTO-ONBOARDING + SELF-SERVICE + SMART AUTOMATION
# ══════════════════════════════════════════════════════════════════════════════

# ── 1. DEALER SELF-SERVICE PORTAL (after activation, no staff needed) ─────────

@app.route("/dealer-portal")
def dealer_portal_redirect():
    """Smart redirect — send dealers to their portal if logged in."""
    t = request.cookies.get("dealer_portal_token","")
    if t in _dealer_portal_sessions:
        did = _dealer_portal_sessions[t]
        return redirect(f"/dealer-portal/{did}/dashboard")
    return redirect("/dealer-portal/login")

_dealer_portal_sessions = {}  # token → dealer_id

@app.route("/dealer-portal/login", methods=["GET","POST"])
def dealer_portal_login():
    if request.method == "GET":
        return render_template("dealer_portal_login.html", dealership=DEALERSHIP_NAME)
    d = request.get_json(silent=True) or request.form
    email = str(d.get("email","")).strip().lower()
    pw    = str(d.get("password",""))
    dealer = next((dd for dd in _dealers.values() if dd.get("email")==email), None)
    if not dealer or not check_password_hash(dealer.get("password_hash",""), pw):
        return jsonify({"ok":False,"message":"Invalid email or password"})
    if dealer.get("status") != "active":
        return jsonify({"ok":False,"message":"Your account is pending activation. We'll email you within 24 hours."})
    token = secrets.token_hex(32)
    _dealer_portal_sessions[token] = dealer["id"]
    resp = jsonify({"ok":True,"redirect":f"/dealer-portal/{dealer['id']}/dashboard"})
    resp.set_cookie("dealer_portal_token", token, max_age=86400*7, httponly=True, samesite="Lax")
    return resp

@app.route("/dealer-portal/logout")
def dealer_portal_logout():
    t = request.cookies.get("dealer_portal_token","")
    _dealer_portal_sessions.pop(t, None)
    resp = redirect("/dealer-portal/login")
    resp.delete_cookie("dealer_portal_token")
    return resp

def dealer_portal_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*a, **kw):
        t = request.cookies.get("dealer_portal_token","")
        if t not in _dealer_portal_sessions:
            return redirect("/dealer-portal/login")
        did = kw.get("dealer_id","") or (a[0] if a else "")
        if _dealer_portal_sessions[t] != str(did):
            return redirect("/dealer-portal/login")
        return f(*a, **kw)
    return decorated

def _current_dealer_id():
    t = request.cookies.get("dealer_portal_token","")
    return _dealer_portal_sessions.get(t)

@app.route("/dealer-portal/<dealer_id>/dashboard")
@dealer_portal_required
def dealer_portal_dashboard(dealer_id):
    dealer = _dealers.get(dealer_id)
    if not dealer: abort(404)
    plan = PLANS.get(dealer.get("plan","free"),{})
    all_v = get_all_vehicles(include_archived=False)
    my_vehicles = [v for v in all_v if v.get("Dealer_ID","") == dealer_id]
    available = [v for v in my_vehicles if v.get("Status")=="Available"]
    sold = [v for v in my_vehicles if v.get("Status")=="Sold"]
    for v in my_vehicles:
        vid = v.get("Internal_ID","").upper()
        v["_primary_photo"] = next((p["url"] for p in _vehicle_photos.get(vid,[]) if p.get("primary")), None)
        v["_photo_count"] = len(_vehicle_photos.get(vid,[]))
    stock_value = sum(_vehicle_price_num(v) for v in available)
    base_url = request.host_url.rstrip("/")
    return render_template("dealer_portal_dashboard.html",
        dealer=dealer, plan=plan, vehicles=my_vehicles,
        available=available, sold=sold,
        stock_value=stock_value, base_url=base_url,
        dealership=DEALERSHIP_NAME)

@app.route("/dealer-portal/<dealer_id>/add-vehicle", methods=["POST"])
@dealer_portal_required
def dealer_portal_add_vehicle(dealer_id):
    dealer = _dealers.get(dealer_id)
    if not dealer: abort(404)
    plan = PLANS.get(dealer.get("plan","free"),{})
    all_v = get_all_vehicles(include_archived=False)
    my_count = sum(1 for v in all_v if v.get("Dealer_ID","")==dealer_id)
    if my_count >= plan.get("max_listings",5):
        return jsonify({"ok":False,"message":f"Your {plan['name']} plan allows {plan['max_listings']} listings. Upgrade to add more."})
    d = request.get_json(silent=True) or {}
    fields = {k:v for k,v in {
        "Internal_ID": d.get("internal_id","").strip().upper() or f"{dealer_id[:3]}{str(uuid.uuid4())[:6].upper()}",
        "Make": d.get("make","").strip(),
        "Model": d.get("model","").strip(),
        "Year": int(d.get("year",0) or 0) or None,
        "Trim": d.get("trim","").strip(),
        "Body": d.get("body","").strip(),
        "Engine": d.get("engine","").strip(),
        "Transmission": d.get("transmission","").strip(),
        "Fuel": d.get("fuel","").strip(),
        "Color": d.get("color","").strip(),
        "Mileage": str(d.get("mileage","")).strip(),
        "Price": _to_num(d.get("price","")),
        "VIN": d.get("vin","").strip(),
        "Notes": d.get("notes","").strip(),
        "Status": "Available",
        "Dealer_ID": dealer_id,
    }.items() if v}
    ok, msg = create_vehicle_record(fields)
    if ok:
        _dealers[dealer_id]["vehicles_count"] = my_count + 1
        log_activity("dealer_add_vehicle", fields["Internal_ID"], dealer["name"], "portal")
        _notify_buyers_new_listing(fields)
    return jsonify({"ok":ok,"message":msg,"internal_id":fields.get("Internal_ID","")})

@app.route("/dealer-portal/<dealer_id>/delete-vehicle", methods=["POST"])
@dealer_portal_required
def dealer_portal_delete_vehicle(dealer_id):
    d = request.get_json(silent=True) or {}
    vid = d.get("vid","").upper()
    v = get_vehicle(vid)
    if not v or v.get("Dealer_ID","") != dealer_id:
        return jsonify({"ok":False,"message":"Vehicle not found or not yours"})
    ok, msg = patch_vehicle(vid, {"Status":"Archived"}, user=f"dealer:{dealer_id}")
    return jsonify({"ok":ok,"message":msg})

@app.route("/dealer-portal/<dealer_id>/profile", methods=["GET","POST"])
@dealer_portal_required
def dealer_portal_profile(dealer_id):
    dealer = _dealers.get(dealer_id)
    if not dealer: abort(404)
    if request.method == "POST":
        d = request.get_json(silent=True) or {}
        for field in ["name","phone","address","suburb","state"]:
            if d.get(field):
                _dealers[dealer_id][field] = str(d[field]).strip()
        return jsonify({"ok":True})
    return render_template("dealer_portal_profile.html",
        dealer=dealer, plans=PLANS, dealership=DEALERSHIP_NAME)

# ── 2. SMART ONBOARDING WIZARD ────────────────────────────────────────────────

_onboarding_state = {}   # session_id → { step, data }

@app.route("/onboard")
def onboard_start():
    """Public smart onboarding — works for both buyers and dealers."""
    return render_template("onboard.html", dealership=DEALERSHIP_NAME)

@app.route("/api/onboard/step", methods=["POST"])
def api_onboard_step():
    d = request.get_json(silent=True) or {}
    user_type = d.get("type","")   # "buyer" or "dealer"
    step = d.get("step", 1)
    data = d.get("data", {})
    sid  = d.get("session_id", secrets.token_hex(16))

    if sid not in _onboarding_state:
        _onboarding_state[sid] = {"step":1,"type":user_type,"data":{}}

    state = _onboarding_state[sid]
    state["data"].update(data)
    state["type"] = user_type or state.get("type","")

    # Buyer onboarding: step 1=name+email+phone+pw → auto register → done
    if state["type"] == "buyer":
        if step == 2:
            email = state["data"].get("email","").strip().lower()
            name  = state["data"].get("name","").strip()
            phone = state["data"].get("phone","").strip()
            pw    = state["data"].get("password","")
            if email in _buyers:
                return jsonify({"ok":False,"message":"Email already registered — sign in instead","redirect":"/buyer/login"})
            if len(pw) < 6:
                return jsonify({"ok":False,"message":"Password must be at least 6 characters"})
            _buyers[email] = {
                "email":email,"name":name,"phone":phone,
                "password_hash":generate_password_hash(pw),
                "created":datetime.now(timezone.utc).isoformat(),
                "saved":[],"alerts":[],"notifications":[],
            }
            token = secrets.token_hex(32)
            _buyer_sessions[token] = email
            resp = jsonify({"ok":True,"done":True,
                "message":f"Welcome {name}! Your account is ready.",
                "redirect":"/buyer/dashboard"})
            resp.set_cookie("buyer_token",token,max_age=86400*30,httponly=True,samesite="Lax")
            return resp
        return jsonify({"ok":True,"session_id":sid,"next_step":2})

    # Dealer onboarding: 3 steps
    if state["type"] == "dealer":
        if step == 4:  # Final submit
            dd = state["data"]
            name  = dd.get("name","").strip()
            email = dd.get("email","").strip().lower()
            pw    = dd.get("password","")
            plan  = dd.get("plan","dealer")
            if not name or not email or not pw:
                return jsonify({"ok":False,"message":"Name, email and password required"})
            if any(x.get("email")==email for x in _dealers.values()):
                return jsonify({"ok":False,"message":"Email already registered"})
            did = _dealer_id()
            _dealers[did] = {
                "id":did,"name":name,"email":email,
                "phone":dd.get("phone","").strip(),
                "address":dd.get("address","").strip(),
                "suburb":dd.get("suburb","").strip(),
                "state":dd.get("state","NSW"),
                "plan":plan if plan in PLANS else "dealer",
                "status":"pending",
                "password_hash":generate_password_hash(pw),
                "airtable_token":dd.get("airtable_token","").strip(),
                "airtable_base":dd.get("airtable_base","").strip(),
                "logo_url":"","created":datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "vehicles_count":0,"subscription_end":"",
                "slug":re.sub(r"[^a-z0-9]+","-",name.lower()).strip("-"),
            }
            _onboarding_state.pop(sid,None)
            send_sms(MANAGER_PHONE,f"NEW DEALER via wizard: {name} ({email}) — {plan} plan. Activate at /platform/admin")
            return jsonify({"ok":True,"done":True,
                "message":"You're in! We'll activate your account within 24 hours.",
                "dealer_id":did})
        return jsonify({"ok":True,"session_id":sid,"next_step":step+1})

    return jsonify({"ok":True,"session_id":sid})


# ── 3. AUTO AI VEHICLE DESCRIPTION ───────────────────────────────────────────

@app.route("/api/dealer-portal/ai-describe", methods=["POST"])
def api_dealer_ai_describe():
    """Auto-generate a vehicle description from specs. No auth — called during listing."""
    d = request.get_json(silent=True) or {}
    specs = " ".join(filter(None,[
        str(d.get("year","")), str(d.get("make","")), str(d.get("model","")),
        str(d.get("trim","")), str(d.get("body","")), str(d.get("engine","")),
        str(d.get("transmission","")), str(d.get("fuel","")), str(d.get("color","")),
        str(d.get("mileage","")), str(d.get("price","")),
    ]))
    if not specs.strip():
        return jsonify({"ok":False,"message":"Enter vehicle details first"})
    if not OPENAI_API_KEY:
        return jsonify({"ok":True,"description":f"Quality {d.get('year','')} {d.get('make','')} {d.get('model','')} in excellent condition. Low kilometres, well maintained, drives beautifully. Contact us today for a test drive."})
    try:
        import openai
        client = openai.OpenAI(api_key=OPENAI_API_KEY)
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role":"user","content":f"Write a compelling 3-sentence car listing description for: {specs}. Focus on key selling points, be enthusiastic but honest. No markdown, plain text only."}],
            max_tokens=150, temperature=0.7
        )
        return jsonify({"ok":True,"description":resp.choices[0].message.content.strip()})
    except Exception as e:
        return jsonify({"ok":False,"message":str(e)[:100]})


# ── 4. AUTO VIN DECODE ON LISTING ─────────────────────────────────────────────

@app.route("/api/dealer-portal/decode-vin", methods=["POST"])
def api_dealer_decode_vin():
    """Decode VIN and auto-fill vehicle fields — no login needed."""
    vin = (request.get_json(silent=True) or {}).get("vin","").strip().upper()
    if len(vin) != 17:
        return jsonify({"ok":False,"message":"VIN must be 17 characters"})
    try:
        r = requests.get(f"https://vpic.nhtsa.dot.gov/api/vehicles/decodevin/{vin}?format=json", timeout=8)
        results = r.json().get("Results",[])
        def get_val(var):
            return next((x["Value"] for x in results if x.get("Variable")==var and x.get("Value") and x["Value"]!="Not Applicable"),"")
        data = {
            "make": get_val("Make").title(),
            "model": get_val("Model"),
            "year": get_val("Model Year"),
            "body": get_val("Body Class"),
            "engine": f"{get_val('Displacement (L)')}L {get_val('Engine Number of Cylinders')}cyl".strip("L cyl").strip() or get_val("Engine Configuration"),
            "transmission": get_val("Transmission Style"),
            "fuel": get_val("Fuel Type - Primary"),
            "trim": get_val("Trim"),
        }
        return jsonify({"ok":True,"data":{k:v for k,v in data.items() if v}})
    except Exception as e:
        return jsonify({"ok":False,"message":"VIN decode failed — fill in manually"})


# ── 5. LIVE CHAT / MESSAGING SYSTEM ──────────────────────────────────────────

_chats = {}   # chat_id → { buyer_email, dealer_id, vid, messages:[], created, unread_dealer }

@app.route("/api/chat/start", methods=["POST"])
def api_chat_start():
    d = request.get_json(silent=True) or {}
    buyer_email = _current_buyer()
    if not buyer_email:
        return jsonify({"ok":False,"login_required":True})
    vid    = d.get("vid","").upper()
    dealer_id = d.get("dealer_id","")
    # Find existing chat
    existing = next((c for c in _chats.values()
                     if c["buyer_email"]==buyer_email and c["vid"]==vid), None)
    if existing:
        return jsonify({"ok":True,"chat_id":existing["id"],"existing":True})
    chat_id = "CHT" + str(uuid.uuid4())[:8].upper()
    vehicle = get_vehicle(vid) if vid else {}
    _chats[chat_id] = {
        "id": chat_id, "buyer_email": buyer_email,
        "dealer_id": dealer_id or "platform",
        "vid": vid, "vehicle_name": f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')}".strip(),
        "messages": [], "created": datetime.now(timezone.utc).isoformat(),
        "unread_dealer": 0, "unread_buyer": 0,
    }
    return jsonify({"ok":True,"chat_id":chat_id,"existing":False})

@app.route("/api/chat/<chat_id>/messages", methods=["GET"])
def api_chat_get(chat_id):
    chat = _chats.get(chat_id)
    if not chat: return jsonify({"ok":False,"message":"Chat not found"})
    buyer = _current_buyer()
    if buyer and buyer == chat["buyer_email"]:
        chat["unread_buyer"] = 0
    return jsonify({"ok":True,"messages":chat["messages"],"chat":{"id":chat_id,"vehicle_name":chat.get("vehicle_name","")}})

@app.route("/api/chat/<chat_id>/send", methods=["POST"])
def api_chat_send(chat_id):
    chat = _chats.get(chat_id)
    if not chat: return jsonify({"ok":False,"message":"Chat not found"})
    d = request.get_json(silent=True) or {}
    text = d.get("message","").strip()
    if not text: return jsonify({"ok":False,"message":"Empty message"})
    sender = d.get("sender","buyer")   # "buyer" or "dealer"
    buyer_email = _current_buyer()
    if sender == "buyer" and not buyer_email:
        return jsonify({"ok":False,"login_required":True})
    msg = {
        "id": str(uuid.uuid4())[:8], "text": text,
        "sender": sender, "time": datetime.now(timezone.utc).strftime("%H:%M"),
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }
    chat["messages"].append(msg)
    if sender == "buyer":
        chat["unread_dealer"] += 1
        # Notify dealer/manager via SMS
        buyer = _buyers.get(buyer_email,{})
        send_sms(MANAGER_PHONE, f"NEW CHAT from {buyer.get('name',buyer_email)} re {chat.get('vehicle_name','')}: {text[:80]}")
    else:
        chat["unread_buyer"] += 1
        # Notify buyer via SMS
        buyer = _buyers.get(chat["buyer_email"],{})
        if buyer.get("phone"):
            send_sms(buyer["phone"], f"Reply from {DEALERSHIP_NAME} re {chat.get('vehicle_name','')}: {text[:80]}")
    return jsonify({"ok":True,"message":msg})

@app.route("/api/chat/dealer-inbox")
@login_required
def api_dealer_chat_inbox():
    """All chats for dealer staff to respond to."""
    inbox = sorted(_chats.values(), key=lambda c: c.get("created",""), reverse=True)
    return jsonify({"ok":True,"chats":inbox,
                    "unread_total":sum(c.get("unread_dealer",0) for c in inbox)})

@app.route("/chat-inbox")
@login_required
def chat_inbox_page():
    return render_template("chat_inbox.html", dealership=DEALERSHIP_NAME)


# ── 6. AUTOMATED FOLLOW-UP ENGINE ────────────────────────────────────────────

_auto_followup_log = []   # list of {type, contact, message, sent_at, vid}

def _send_auto_followup(buyer_email, vid, trigger):
    """Auto SMS/notification based on buyer behaviour."""
    buyer = _buyers.get(buyer_email,{})
    if not buyer: return
    vehicle = get_vehicle(vid) if vid else {}
    vname = f"{vehicle.get('Year','')} {vehicle.get('Make','')} {vehicle.get('Model','')}".strip() or vid

    messages = {
        "saved_3days": f"Hi {buyer.get('name','there')}! Still interested in the {vname}? It's still available at {DEALERSHIP_NAME}. Reply to book a test drive.",
        "viewed_twice": f"The {vname} you've been looking at won't last long. Want to come in for a test drive? Call us or reply here.",
        "price_drop_saved": f"Good news! The {vname} you saved just dropped in price. Now more affordable than ever.",
        "enquiry_no_response": f"Hi {buyer.get('name','there')}, following up on your enquiry about the {vname}. Have you had a chance to consider it?",
    }
    msg = messages.get(trigger,"")
    if not msg: return

    note = {"type":trigger,"contact":buyer_email,
            "message":msg,"sent_at":datetime.now(timezone.utc).strftime("%b %d %H:%M"),"vid":vid}
    _auto_followup_log.insert(0,note)

    # Push as notification
    buyer.setdefault("notifications",[]).insert(0,{
        "id":str(uuid.uuid4())[:8],"type":"followup",
        "message":msg,"vid":vid,"vname":vname,
        "date":datetime.now(timezone.utc).strftime("%b %d %H:%M"),"read":False,
    })
    if buyer.get("phone"):
        send_sms(buyer["phone"], msg)

@app.route("/api/auto-followup/log")
@login_required
def api_followup_log():
    return jsonify({"ok":True,"log":_auto_followup_log[:50]})


# ── 7. REVIEWS & RATINGS ──────────────────────────────────────────────────────

_reviews = []   # { id, buyer_email, buyer_name, dealer_id, vid, rating, text, date, approved }

@app.route("/api/reviews/add", methods=["POST"])
def api_add_review():
    buyer_email = _current_buyer()
    if not buyer_email:
        return jsonify({"ok":False,"login_required":True})
    buyer = _buyers.get(buyer_email,{})
    d = request.get_json(silent=True) or {}
    rating = int(d.get("rating",5))
    if not 1 <= rating <= 5:
        return jsonify({"ok":False,"message":"Rating must be 1-5"})
    review = {
        "id": str(uuid.uuid4())[:8],
        "buyer_email": buyer_email,
        "buyer_name": buyer.get("name","Anonymous"),
        "dealer_id": d.get("dealer_id",""),
        "vid": d.get("vid",""),
        "rating": rating,
        "text": str(d.get("text","")).strip()[:500],
        "date": datetime.now(timezone.utc).strftime("%b %d %Y"),
        "approved": False,   # Admin approves before showing
    }
    _reviews.insert(0, review)
    log_activity("review_added", d.get("vid",""), f"{rating}★ — {buyer.get('name','')}", buyer_email)
    return jsonify({"ok":True,"message":"Thank you for your review! It will appear after moderation."})

@app.route("/api/reviews/<dealer_id>")
def api_get_reviews(dealer_id):
    approved = [r for r in _reviews if r.get("dealer_id")==dealer_id and r.get("approved")]
    avg = round(sum(r["rating"] for r in approved)/len(approved),1) if approved else 0
    return jsonify({"ok":True,"reviews":approved,"average":avg,"total":len(approved)})

@app.route("/api/reviews/approve", methods=["POST"])
@login_required
def api_approve_review():
    rid = (request.get_json(silent=True) or {}).get("id","")
    for r in _reviews:
        if r["id"] == rid:
            r["approved"] = True
            return jsonify({"ok":True})
    return jsonify({"ok":False})


# ── 8. SMART HOMEPAGE ─────────────────────────────────────────────────────────

@app.route("/home")
def smart_home():
    """Personalised homepage — shows relevant content based on buyer history."""
    buyer_email = _current_buyer()
    buyer = _buyers.get(buyer_email,{}) if buyer_email else {}
    all_v = get_all_vehicles(include_archived=False)
    available = [v for v in all_v if v.get("Status")=="Available"]
    # Featured = most recently added 6
    featured = sorted(available, key=lambda v:str(v.get("Internal_ID","")), reverse=True)[:6]
    # Price drops
    drops = []
    for vid, history in _price_history.items():
        if len(history)>=2:
            try:
                curr=float(re.sub(r"[^\d.]","",str(history[-1]["price"]) or "")or 0)
                prev=float(re.sub(r"[^\d.]","",str(history[-2]["price"]) or "")or 0)
                if curr < prev:
                    v=get_vehicle(vid)
                    if v and v.get("Status")=="Available":
                        v["_drop_amount"]=round(prev-curr)
                        v["_primary_photo"]=next((p["url"] for p in _vehicle_photos.get(vid,[]) if p.get("primary")),None)
                        drops.append(v)
            except: pass
    # Personalised picks (based on saved vehicles' makes)
    picks = []
    if buyer.get("saved"):
        saved_makes = set()
        for vid in buyer["saved"]:
            v=get_vehicle(vid)
            if v: saved_makes.add(v.get("Make","").lower())
        picks = [v for v in available if v.get("Make","").lower() in saved_makes
                 and v.get("Internal_ID","") not in buyer["saved"]][:4]
    for v in featured+drops[:4]+picks:
        vid=v.get("Internal_ID","").upper()
        if "_primary_photo" not in v:
            v["_primary_photo"]=next((p["url"] for p in _vehicle_photos.get(vid,[]) if p.get("primary")),None)
    unread=sum(1 for n in buyer.get("notifications",[]) if not n.get("read")) if buyer else 0
    makes=sorted(set(v.get("Make","") for v in available if v.get("Make")))
    bodies=sorted(set(v.get("Body","") for v in available if v.get("Body")))
    return render_template("smart_home.html",
        dealership=DEALERSHIP_NAME, buyer=buyer,
        featured=featured, drops=drops[:4], picks=picks,
        total_available=len(available), makes=makes, bodies=bodies,
        unread=unread, dealer_count=len([d for d in _dealers.values() if d.get("status")=="active"]))


# ── 9. SELF-SERVICE UPGRADE PAGE ─────────────────────────────────────────────

@app.route("/dealer-portal/<dealer_id>/upgrade")
@dealer_portal_required
def dealer_portal_upgrade(dealer_id):
    dealer = _dealers.get(dealer_id)
    if not dealer: abort(404)
    return render_template("dealer_portal_upgrade.html",
        dealer=dealer, plans=PLANS, dealership=DEALERSHIP_NAME)

@app.route("/api/dealer-portal/<dealer_id>/request-upgrade", methods=["POST"])
@dealer_portal_required
def api_dealer_request_upgrade(dealer_id):
    dealer = _dealers.get(dealer_id)
    if not dealer: abort(404)
    new_plan = (request.get_json(silent=True) or {}).get("plan","dealer")
    if new_plan not in PLANS: return jsonify({"ok":False,"message":"Invalid plan"})
    send_sms(MANAGER_PHONE, f"UPGRADE REQUEST: {dealer['name']} wants to upgrade to {new_plan} plan. Approve at /platform/admin")
    log_activity("upgrade_request", dealer_id, f"{dealer['plan']} → {new_plan}", "portal")
    return jsonify({"ok":True,"message":"Upgrade request sent! We'll activate within a few hours."})



# ══════════════════════════════════════════════════════════════════════════════
# STRIPE BILLING — Auto-charge dealers on signup & monthly renewals
# ══════════════════════════════════════════════════════════════════════════════

STRIPE_SECRET_KEY      = os.environ.get("STRIPE_SECRET_KEY", "")
STRIPE_PUBLISHABLE_KEY = os.environ.get("STRIPE_PUBLISHABLE_KEY", "")
STRIPE_WEBHOOK_SECRET  = os.environ.get("STRIPE_WEBHOOK_SECRET", "")

# Stripe Price IDs — create these in your Stripe dashboard and paste them here
STRIPE_PRICE_IDS = {
    "dealer":     os.environ.get("STRIPE_PRICE_DEALER",     ""),
    "pro":        os.environ.get("STRIPE_PRICE_PRO",        ""),
    "enterprise": os.environ.get("STRIPE_PRICE_ENTERPRISE", ""),
}

def _stripe_post(endpoint, payload):
    """Minimal Stripe API caller — no SDK needed."""
    if not STRIPE_SECRET_KEY:
        return None, "STRIPE_SECRET_KEY not configured"
    try:
        import base64 as _b64
        auth = _b64.b64encode(f"{STRIPE_SECRET_KEY}:".encode()).decode()
        r = requests.post(
            f"https://api.stripe.com/v1/{endpoint}",
            data=payload,
            headers={"Authorization": f"Basic {auth}",
                     "Content-Type": "application/x-www-form-urlencoded"},
            timeout=15
        )
        data = r.json()
        if "error" in data:
            return None, data["error"].get("message", "Stripe error")
        return data, None
    except Exception as e:
        return None, str(e)

def _stripe_get(endpoint):
    if not STRIPE_SECRET_KEY:
        return None, "STRIPE_SECRET_KEY not configured"
    try:
        import base64 as _b64
        auth = _b64.b64encode(f"{STRIPE_SECRET_KEY}:".encode()).decode()
        r = requests.get(
            f"https://api.stripe.com/v1/{endpoint}",
            headers={"Authorization": f"Basic {auth}"},
            timeout=15
        )
        data = r.json()
        if "error" in data:
            return None, data["error"].get("message", "Stripe error")
        return data, None
    except Exception as e:
        return None, str(e)

# ── Checkout session — dealer pays online ─────────────────────────────────────
@app.route("/api/stripe/create-checkout", methods=["POST"])
def api_stripe_create_checkout():
    d = request.get_json(silent=True) or {}
    dealer_id = d.get("dealer_id", "")
    plan = d.get("plan", "dealer")
    dealer = _dealers.get(dealer_id)
    if not dealer:
        return jsonify({"ok": False, "message": "Dealer not found"})
    price_id = STRIPE_PRICE_IDS.get(plan, "")
    if not price_id:
        return jsonify({"ok": False, "message": f"No Stripe price configured for {plan} plan. Add STRIPE_PRICE_{plan.upper()} env var."})
    base = request.host_url.rstrip("/")
    payload = {
        "mode": "subscription",
        "line_items[0][price]": price_id,
        "line_items[0][quantity]": "1",
        "success_url": f"{base}/dealer-portal/{dealer_id}/dashboard?payment=success",
        "cancel_url": f"{base}/dealer-portal/{dealer_id}/upgrade?payment=cancelled",
        "customer_email": dealer["email"],
        "metadata[dealer_id]": dealer_id,
        "metadata[plan]": plan,
        "allow_promotion_codes": "true",
    }
    session_data, err = _stripe_post("checkout/sessions", payload)
    if err:
        return jsonify({"ok": False, "message": err})
    return jsonify({"ok": True, "url": session_data.get("url", "")})

# ── Stripe webhook — auto-activate on payment ─────────────────────────────────
@app.route("/stripe/webhook", methods=["POST"])
def stripe_webhook():
    payload = request.get_data()
    sig = request.headers.get("Stripe-Signature", "")
    # Verify signature
    if STRIPE_WEBHOOK_SECRET:
        try:
            import hmac as _hmac, hashlib as _hl
            parts = {p.split("=")[0]: p.split("=")[1] for p in sig.split(",") if "=" in p}
            ts = parts.get("t", "")
            v1 = parts.get("v1", "")
            signed = f"{ts}.{payload.decode()}"
            expected = _hmac.new(STRIPE_WEBHOOK_SECRET.encode(), signed.encode(), _hl.sha256).hexdigest()
            if not _hmac.compare_digest(expected, v1):
                return "Invalid signature", 400
        except Exception:
            return "Signature check failed", 400

    try:
        event = json.loads(payload)
    except Exception:
        return "Bad JSON", 400

    etype = event.get("type", "")

    if etype == "checkout.session.completed":
        meta = event["data"]["object"].get("metadata", {})
        dealer_id = meta.get("dealer_id", "")
        plan = meta.get("plan", "dealer")
        sub_id = event["data"]["object"].get("subscription", "")
        if dealer_id and dealer_id in _dealers:
            _dealers[dealer_id]["status"] = "active"
            _dealers[dealer_id]["plan"] = plan
            _dealers[dealer_id]["stripe_subscription_id"] = sub_id
            _dealers[dealer_id]["subscription_end"] = ""
            log_activity("stripe_payment", dealer_id, f"Activated {plan} via Stripe", "stripe")
            send_sms(MANAGER_PHONE, f"💳 PAYMENT: {_dealers[dealer_id]['name']} paid for {plan} plan. Auto-activated.")

    elif etype in ("customer.subscription.deleted", "invoice.payment_failed"):
        sub_id = event["data"]["object"].get("id") or event["data"]["object"].get("subscription", "")
        for dealer in _dealers.values():
            if dealer.get("stripe_subscription_id") == sub_id:
                dealer["status"] = "suspended"
                log_activity("stripe_suspended", dealer["id"], "Payment failed / cancelled", "stripe")
                break

    elif etype == "invoice.payment_succeeded":
        sub_id = event["data"]["object"].get("subscription", "")
        for dealer in _dealers.values():
            if dealer.get("stripe_subscription_id") == sub_id:
                dealer["status"] = "active"
                log_activity("stripe_renewed", dealer["id"], "Subscription renewed", "stripe")
                break

    return "ok", 200

# ── Portal: billing info ───────────────────────────────────────────────────────
@app.route("/api/dealer-portal/<dealer_id>/billing-status")
@dealer_portal_required
def api_dealer_billing_status(dealer_id):
    dealer = _dealers.get(dealer_id)
    if not dealer: abort(404)
    sub_id = dealer.get("stripe_subscription_id", "")
    sub_data = None
    if sub_id:
        sub_data, _ = _stripe_get(f"subscriptions/{sub_id}")
    return jsonify({
        "ok": True,
        "plan": dealer.get("plan", "free"),
        "status": dealer.get("status", "pending"),
        "stripe_connected": bool(sub_id),
        "subscription": sub_data,
        "publishable_key": STRIPE_PUBLISHABLE_KEY,
    })

# ── Manage subscription portal (Stripe customer portal) ───────────────────────
@app.route("/api/dealer-portal/<dealer_id>/manage-billing", methods=["POST"])
@dealer_portal_required
def api_dealer_manage_billing(dealer_id):
    dealer = _dealers.get(dealer_id)
    if not dealer: abort(404)
    # Need customer ID from subscription
    sub_id = dealer.get("stripe_subscription_id", "")
    if not sub_id:
        return jsonify({"ok": False, "message": "No active subscription found."})
    sub_data, err = _stripe_get(f"subscriptions/{sub_id}")
    if err or not sub_data:
        return jsonify({"ok": False, "message": "Could not retrieve subscription."})
    cust_id = sub_data.get("customer", "")
    base = request.host_url.rstrip("/")
    portal, err = _stripe_post("billing_portal/sessions", {
        "customer": cust_id,
        "return_url": f"{base}/dealer-portal/{dealer_id}/upgrade",
    })
    if err:
        return jsonify({"ok": False, "message": err})
    return jsonify({"ok": True, "url": portal.get("url", "")})



# ══════════════════════════════════════════════════════════════════════════════
# PWA — Progressive Web App routes
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/manifest.json")
def pwa_manifest():
    from flask import send_from_directory
    return send_from_directory("static", "manifest.json",
                               mimetype="application/manifest+json")

@app.route("/sw.js")
def pwa_sw():
    from flask import send_from_directory, make_response
    resp = make_response(send_from_directory("static", "sw.js",
                                             mimetype="application/javascript"))
    resp.headers["Service-Worker-Allowed"] = "/"
    resp.headers["Cache-Control"] = "no-cache"
    return resp

@app.route("/offline")
def pwa_offline():
    return render_template("pwa_offline.html", dealership=DEALERSHIP_NAME)

@app.route("/static/icon-<int:size>.png")
def pwa_icon(size):
    """Generate a simple SVG-based PNG icon on the fly."""
    try:
        # Try to serve a real file if it exists
        from flask import send_from_directory
        return send_from_directory("static", f"icon-{size}.png")
    except Exception:
        pass
    # Generate a minimal 1x1 gold PNG as fallback
    import base64
    # 1x1 gold pixel PNG
    gold_px = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwADhQGAWjR9awAAAABJRU5ErkJggg=="
    )
    from flask import Response
    return Response(gold_px, mimetype="image/png",
                    headers={"Cache-Control": "public, max-age=86400"})

# Inject PWA tags into all public pages via a template context processor
@app.context_processor
def inject_pwa():
    return {
        "pwa_enabled": True,
        "pwa_name": DEALERSHIP_NAME,
    }



# ══════════════════════════════════════════════════════════════════════════════
# STAFF PHONE / CONTACT MANAGEMENT
# Each staff member has their own phone number stored in _staff_accounts.
# When they share a link (/share/<id>?ref=<username>) the customer sees THEIR
# number. Admin can update any staff phone at any time from Admin → Staff.
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/staff/update-contact", methods=["POST"])
@admin_required
def admin_staff_update_contact():
    """Admin edits any staff member's phone or email — live, instant effect."""
    d = request.get_json(silent=True) or {}
    username = d.get("username", "").strip().lower()
    if username not in _staff_accounts:
        return jsonify({"ok": False, "message": "Staff account not found"})
    phone = d.get("phone", "").strip()
    email = d.get("email", "").strip()
    name  = d.get("name",  "").strip()
    role  = d.get("role",  "").strip()
    if phone is not None:
        _staff_accounts[username]["phone"] = phone
    if email is not None:
        _staff_accounts[username]["email"] = email
    if name:
        _staff_accounts[username]["name"] = name
    if role and role in ("admin", "manager", "sales", "service", "viewer", "staff"):
        _staff_accounts[username]["role"] = role
    log_activity("staff_update", "", f"Updated contact for '{username}': phone={phone}", "admin")
    return jsonify({"ok": True, "message": f"Contact updated for {username}"})

@app.route("/api/staff/my-contact", methods=["POST"])
@login_required
def api_staff_update_own_contact():
    """Staff member updates their own phone/email from dashboard."""
    d = request.get_json(silent=True) or {}
    username = session.get("username", "")
    if not username or username not in _staff_accounts:
        return jsonify({"ok": False, "message": "Not found"})
    phone = d.get("phone", "").strip()
    email = d.get("email", "").strip()
    if phone:
        _staff_accounts[username]["phone"] = phone
    if email:
        _staff_accounts[username]["email"] = email
    return jsonify({"ok": True})

